# -*- encoding: utf-8 -*-
"""样品单交期执行看板。

页面只持久化人工录入字段；交样周期、剩余工作日、预期状态、考核天数和考核分数均在读取时
实时计算，避免日期变化后数据库中保存的派生值失真。
"""

import copy
import io
import logging
import re
import time
import uuid
import warnings
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from html import escape
from typing import Any, Optional
from urllib.parse import quote

from chinese_calendar import is_holiday
from nicegui import app, ui
from nicegui.events import UploadEventArguments
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from .. import db_storage
from ..config import IMG_DIR, PRESET_AVATARS
from ..issue_workflow_utils import merge_wecom_recipients, schedule_background_task
from ..sample_order_dashboard_config import (
    SAMPLE_ORDER_ADMIN_ROLES,
    SAMPLE_ORDER_BASE_EDITOR_ROLES,
    SAMPLE_ORDER_DELAY_ATTENTION_THRESHOLD,
    SAMPLE_ORDER_DELAY_EDITOR_ROLES,
    SAMPLE_ORDER_DELAY_NATURE_MARKER_ROLES,
    SAMPLE_ORDER_MANAGER_NOTIFY_TARGETS,
    SAMPLE_ORDER_NOTIFY_APPLICANT_ON_EXTENSION,
    SAMPLE_ORDER_NOTIFY_APPLICANT_ON_SPECIAL_STATUS,
    SAMPLE_ORDER_PUBLIC_BASE_URL,
    SAMPLE_ORDER_REDIRECT_APPLICANT_NOTIFICATIONS_TO_MANAGER,
    SAMPLE_ORDER_SPECIAL_STATUS_EDITOR_ROLES,
    SAMPLE_ORDER_SPECIAL_STATUS_REASON_REQUIRED,
    SAMPLE_ORDER_SPECIAL_STATUSES,
    SAMPLE_ORDER_WARNING_DAYS,
)
from ..utils import get_cache_busted_path, logout, setup_global_activity_tracking, sync_current_user_role
from ..wecom_service import find_unknown_wecom_names, resolve_wecom_recipients, send_wecom_text_message

SAMPLE_ORDER_DATA_KEY = "sample_order_dashboard_data"
SAMPLE_ORDER_ENTITY_NAMESPACE = "sample_order_dashboard"
SAMPLE_ORDER_VERSION_KEY = "sample_order_dashboard_version_stamp"
SAMPLE_ORDER_EXCEL_IMPORT_OWNER = "叶子浩"

logger = logging.getLogger(__name__)


async def initialize_sample_order_storage() -> int:
    """启动时把旧版整块JSON数据安全迁移到逐订单实体表。"""
    migrated_count = await db_storage.migrate_json_dict_to_entities(
        SAMPLE_ORDER_ENTITY_NAMESPACE,
        SAMPLE_ORDER_DATA_KEY,
    )
    if not db_storage.is_json_entity_namespace_initialized(SAMPLE_ORDER_ENTITY_NAMESPACE):
        raise RuntimeError("样品单实体存储初始化失败，已阻止系统以空数据状态启动")
    return migrated_count


def get_all_sample_order_records() -> dict[str, Any]:
    """从逐订单实体缓存读取全部样品单的安全副本。"""
    return db_storage.get_json_entities(SAMPLE_ORDER_ENTITY_NAMESPACE)


FILTER_ALL = "全部"
FILTER_IN_PROGRESS = "制样中"
FILTER_COMPLETED = "已完成"
FILTER_WARNING = "预警"
FILTER_DELAYED = "延期"
FILTER_MANY_DELAYS = "多次延期"
FILTER_NATURE_PENDING = "待性质标记"
DEFAULT_SAMPLE_ORDER_FILTER = FILTER_IN_PROGRESS
SAMPLE_ORDER_CARD_PAGE_SIZE = 30
FILTER_OPTIONS = list(
    dict.fromkeys(
        [
            FILTER_ALL,
            FILTER_IN_PROGRESS,
            FILTER_COMPLETED,
            FILTER_WARNING,
            FILTER_DELAYED,
            FILTER_MANY_DELAYS,
            FILTER_NATURE_PENDING,
            *(status for status in SAMPLE_ORDER_SPECIAL_STATUSES if status != "正常"),
        ]
    )
)


@dataclass
class SampleOrderUpdateResult:
    """一次样品单写操作的结构化结果。"""

    db_success: bool
    changed: bool
    code: str
    record: Optional[dict] = None
    notification_failures: tuple[str, ...] = ()


@dataclass
class SampleOrderImportPreview:
    """Excel解析后的样品单导入预览。"""

    records: list[dict]
    errors: list[str]
    warnings: list[str]
    source_name: str
    total_rows: int


@dataclass
class SampleOrderImportResult:
    """一次Excel批量导入的结构化结果。"""

    db_success: bool
    imported_count: int
    code: str


def option_text(value: object, default: str = "") -> str:
    """把 NiceGUI 宽类型值收窄为可用于业务判断的字符串。"""
    return value.strip() if isinstance(value, str) else default


def option_text_in(value: object, allowed_values: list[str], default: str) -> str:
    """把选择框宽类型值收窄到指定选项集合。"""
    normalized = option_text(value)
    return normalized if normalized in allowed_values else default


def sample_order_delivery_display(value: object) -> str:
    """把空实际交样日期显示为明确的未交样状态。"""
    return option_text(value) or "未交样"


def normalize_int(value: object, default: int = 0) -> int:
    """把输入组件值安全转换为整数。"""
    if isinstance(value, bool):
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        try:
            return int(float(value.strip()))
        except ValueError:
            return default
    return default


def parse_iso_date(value: object) -> Optional[date]:
    """解析存储中的 ISO 日期字符串，并兼容 date/datetime 测试输入。"""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str) or not value.strip():
        return None
    try:
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def is_business_day(day: date) -> bool:
    """按工作日口径判断日期：周末和中国法定节假日均排除。"""
    if day.weekday() >= 5:
        return False
    try:
        return not is_holiday(day)
    except (NotImplementedError, ValueError):
        # 日历库尚未覆盖的年份仍按周一至周五计算，避免页面整体不可用。
        return True


def business_days_between(start: date, end: date) -> int:
    """计算两个日期之间的工作日差，不计起始日，计入结束日。"""
    if start == end:
        return 0
    if end > start:
        cursor = start + timedelta(days=1)
        count = 0
        while cursor <= end:
            if is_business_day(cursor):
                count += 1
            cursor += timedelta(days=1)
        return count
    return -business_days_between(end, start)


def get_sample_order_template() -> dict:
    """返回一张空白样品单的可持久化结构。"""
    return {
        "record_id": "",
        "basic_info": {
            "sample_order_no": "",
            "customer_code": "",
            "product_model": "",
            "application_qty": 1,
            "application_date": "",
            "applicant": "",
            "planned_delivery_date": "",
            "remark": "",
        },
        "execution": {
            "actual_delivery_date": "",
            "sample_owner": "",
        },
        "extensions": [],
        "special_status": {
            "status": "正常",
            "reason": "",
            "updated_by": "",
            "updated_role": "",
            "updated_at": "",
            "history": [],
        },
        "delay_nature": {
            "tag": "",
            "marked_by": "",
            "marked_role": "",
            "marked_at": "",
            "history": [],
        },
        "import_info": {
            "source_name": "",
            "source_row": 0,
        },
        "created_by": "",
        "created_role": "",
        "created_at": "",
        "updated_by": "",
        "updated_role": "",
        "updated_at": "",
        "_revision": 0,
        "operation_log": [],
    }


def merge_with_sample_order_template(raw: object) -> dict:
    """把旧记录或不完整记录补齐为当前数据结构。"""
    merged = get_sample_order_template()
    if not isinstance(raw, dict):
        return merged
    for key in (
        "record_id",
        "created_by",
        "created_role",
        "created_at",
        "updated_by",
        "updated_role",
        "updated_at",
        "_revision",
    ):
        if key in raw:
            merged[key] = copy.deepcopy(raw[key])
    for section in (
        "basic_info",
        "execution",
        "special_status",
        "delay_nature",
        "import_info",
    ):
        source = raw.get(section)
        if isinstance(source, dict):
            merged[section].update(copy.deepcopy(source))
    extensions = raw.get("extensions")
    if isinstance(extensions, list):
        merged["extensions"] = [normalize_extension(item) for item in extensions if isinstance(item, dict)]
    elif isinstance(raw.get("delay"), dict):
        # 兼容首版页面中固定两次延期的数据，迁移后仍只作为历史记录展示。
        legacy_delay = raw["delay"]
        for index, (date_key, reason_key) in enumerate(
            (("first_target_date", "first_reason"), ("second_target_date", "second_reason")),
            start=1,
        ):
            target_date = option_text(legacy_delay.get(date_key))
            reason = option_text(legacy_delay.get(reason_key))
            if target_date or reason:
                merged["extensions"].append(
                    normalize_extension(
                        {
                            "extension_id": f"legacy-{index}",
                            "target_date": target_date,
                            "reason": reason,
                        }
                    )
                )
    operation_log = raw.get("operation_log")
    if isinstance(operation_log, list):
        merged["operation_log"] = copy.deepcopy(operation_log)
    merged["basic_info"]["application_qty"] = max(0, normalize_int(merged["basic_info"].get("application_qty"), 0))
    special_status = merged["special_status"]
    special_status["status"] = option_text_in(
        special_status.get("status"),
        SAMPLE_ORDER_SPECIAL_STATUSES,
        "正常",
    )
    if not isinstance(special_status.get("history"), list):
        special_status["history"] = []
    delay_nature = merged["delay_nature"]
    delay_nature["tag"] = option_text(delay_nature.get("tag"))
    if not isinstance(delay_nature.get("history"), list):
        delay_nature["history"] = []
    merged["_revision"] = max(0, normalize_int(merged.get("_revision"), 0))
    return merged


def normalize_extension(raw: object) -> dict:
    """标准化一条延期历史或尚未保存的延期草稿。"""
    source = raw if isinstance(raw, dict) else {}
    return {
        "extension_id": option_text(source.get("extension_id")),
        "target_date": option_text(source.get("target_date")),
        "reason": option_text(source.get("reason")),
        "created_by": option_text(source.get("created_by")),
        "created_role": option_text(source.get("created_role")),
        "created_at": option_text(source.get("created_at")),
    }


SAMPLE_ORDER_EXCEL_HEADERS = {
    "sample_order_no": "样品单号",
    "customer_code": "客户编码",
    "product_model": "产品型号",
    "application_qty": "申请数量",
    "application_date": "申请日期",
    "applicant": "申请人",
    "planned_delivery_date": "计划交货日期",
    "remark": "备注",
    "actual_delivery_date": "实际交货日期",
    "sample_owner": "制样负责人",
    "first_target_date": "首次延期目标日期",
    "first_reason": "首次延期原因",
    "second_target_date": "二次延期目标日期",
    "second_reason": "二次延期原因",
}


def normalize_sample_order_excel_header(value: object) -> str:
    """移除Excel表头中的换行和空格，便于稳定匹配列。"""
    if not isinstance(value, str):
        return ""
    return re.sub(r"\s+", "", value)


def sample_order_excel_text(value: object) -> str:
    """把Excel单元格转换为不带多余小数位的文本。"""
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value).strip()
    return str(value).strip()


def sample_order_excel_date(value: object) -> str:
    """把Excel日期单元格标准化为ISO日期字符串。"""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.date().isoformat()
            if isinstance(converted, date):
                return converted.isoformat()
            return ""
        except (TypeError, ValueError, OverflowError):
            return ""
    text_value = sample_order_excel_text(value)
    if not text_value:
        return ""
    for date_format in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text_value, date_format).date().isoformat()
        except ValueError:
            continue
    return ""


def parse_sample_order_excel(content: bytes, source_name: str = "") -> SampleOrderImportPreview:
    """解析样品单Excel，忽略物料、支援、公式提示和考核派生列。"""
    if not content:
        return SampleOrderImportPreview([], ["上传文件为空"], [], source_name, 0)
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    except Exception as exc:
        logger.warning("样品单Excel解析失败：%s", exc)
        return SampleOrderImportPreview([], ["文件不是可读取的.xlsx工作簿"], [], source_name, 0)

    worksheet = None
    header_row = 0
    column_map: dict[str, int] = {}
    expected_headers = {
        normalized: key
        for key, header in SAMPLE_ORDER_EXCEL_HEADERS.items()
        if (normalized := normalize_sample_order_excel_header(header))
    }
    for candidate in workbook.worksheets:
        for row_number in range(1, min(candidate.max_row, 10) + 1):
            candidate_map: dict[str, int] = {}
            for column_number, cell in enumerate(candidate[row_number], start=1):
                header_key = expected_headers.get(normalize_sample_order_excel_header(cell.value))
                if header_key:
                    candidate_map[header_key] = column_number
            required_keys = {
                "sample_order_no",
                "customer_code",
                "product_model",
                "application_qty",
                "application_date",
                "applicant",
                "planned_delivery_date",
            }
            if required_keys.issubset(candidate_map):
                worksheet = candidate
                header_row = row_number
                column_map = candidate_map
                break
        if worksheet is not None:
            break
    if worksheet is None:
        workbook.close()
        return SampleOrderImportPreview(
            [],
            ["未找到包含样品单号、客户编码、产品型号等必要表头的工作表"],
            [],
            source_name,
            0,
        )

    records: list[dict] = []
    errors: list[str] = []
    import_warnings: list[str] = []
    total_rows = 0
    for row_number, row_values in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        raw_values = {
            key: row_values[column_number - 1] if column_number <= len(row_values) else None
            for key, column_number in column_map.items()
        }
        if not any(value not in (None, "") for value in raw_values.values()):
            continue
        total_rows += 1
        record = get_sample_order_template()
        basic = record["basic_info"]
        execution = record["execution"]
        basic.update(
            {
                "sample_order_no": sample_order_excel_text(raw_values.get("sample_order_no")),
                "customer_code": sample_order_excel_text(raw_values.get("customer_code")),
                "product_model": sample_order_excel_text(raw_values.get("product_model")),
                "application_qty": normalize_int(raw_values.get("application_qty"), 0),
                "application_date": sample_order_excel_date(raw_values.get("application_date")),
                "applicant": sample_order_excel_text(raw_values.get("applicant")),
                "planned_delivery_date": sample_order_excel_date(raw_values.get("planned_delivery_date")),
                "remark": sample_order_excel_text(raw_values.get("remark")),
            }
        )
        execution.update(
            {
                "actual_delivery_date": sample_order_excel_date(raw_values.get("actual_delivery_date")),
                "sample_owner": SAMPLE_ORDER_EXCEL_IMPORT_OWNER,
            }
        )
        row_warnings: list[str] = []
        for extension_number, (target_key, reason_key) in enumerate(
            (
                ("first_target_date", "first_reason"),
                ("second_target_date", "second_reason"),
            ),
            start=1,
        ):
            raw_target = raw_values.get(target_key)
            target_date = sample_order_excel_date(raw_target)
            reason = sample_order_excel_text(raw_values.get(reason_key))
            if raw_target not in (None, "") and not target_date:
                row_warnings.append(f"第{extension_number}次延期目标日期格式不正确，已忽略该次延期")
            if target_date or reason:
                if target_date and not reason:
                    reason = "历史Excel未填写延期原因"
                    row_warnings.append(f"第{extension_number}次延期缺少原因，已使用占位说明")
                if target_date and reason:
                    record["extensions"].append(normalize_extension({"target_date": target_date, "reason": reason}))
                elif reason:
                    existing_remark = option_text(basic.get("remark"))
                    extra_remark = f"历史Excel第{extension_number}次延期原因：{reason}（目标日期缺失）"
                    basic["remark"] = "；".join(text for text in (existing_remark, extra_remark) if text)
                    row_warnings.append(f"第{extension_number}次延期缺少目标日期，原因已追加到备注")
        validation_errors = validate_sample_order_submission(
            record,
            check_basic=True,
            check_execution=True,
            check_delay=False,
            check_special_status=False,
        )
        row_errors = list(dict.fromkeys(validation_errors))
        if row_errors:
            errors.append(f"第{row_number}行：{'；'.join(row_errors)}")
            continue
        if row_warnings:
            import_warnings.append(f"第{row_number}行：{'；'.join(dict.fromkeys(row_warnings))}")
        record["import_info"].update(
            {
                "source_name": source_name,
                "source_row": row_number,
            }
        )
        records.append(record)
    workbook.close()
    return SampleOrderImportPreview(records, errors, import_warnings, source_name, total_rows)


def _role_matches(role: object, allowed_roles: list[str]) -> bool:
    role_text = option_text(role).lower()
    return any(allowed.lower() in role_text for allowed in allowed_roles)


def is_sample_order_admin(role: object) -> bool:
    """判断角色是否拥有样品单模块管理权限。"""
    return _role_matches(role, SAMPLE_ORDER_ADMIN_ROLES)


def is_sample_order_base_editor(role: object) -> bool:
    """判断角色是否可以录入和维护基础信息。"""
    return is_sample_order_admin(role) or _role_matches(role, SAMPLE_ORDER_BASE_EDITOR_ROLES)


def is_sample_order_delay_editor(role: object) -> bool:
    """判断角色是否可以维护制样执行与延期信息。"""
    return is_sample_order_admin(role) or _role_matches(role, SAMPLE_ORDER_DELAY_EDITOR_ROLES)


def is_sample_order_special_status_editor(role: object) -> bool:
    """判断角色是否可以设置暂停、作废等特殊状态。"""
    return is_sample_order_admin(role) or _role_matches(role, SAMPLE_ORDER_SPECIAL_STATUS_EDITOR_ROLES)


def is_sample_order_delay_nature_marker(role: object) -> bool:
    """判断角色是否可以为已完成延期订单标记性质。"""
    return is_sample_order_admin(role) or _role_matches(role, SAMPLE_ORDER_DELAY_NATURE_MARKER_ROLES)


def can_view_sample_order_average_score(role: object) -> bool:
    """平均考核分仅向研发样品组长和研发经理展示。"""
    return _role_matches(role, ["研发样品组长", "研发经理"])


def sample_order_matches_kpi(record: object, metrics: dict, label: str) -> bool:
    """判断一张样品单是否属于指定的顶部统计卡片。"""
    if label == "制样中":
        return sample_order_matches_filter(
            record,
            FILTER_IN_PROGRESS,
            calculated_metrics=metrics,
        )
    if label == "预警":
        return metrics.get("attention_level") in {"missing", "warning"}
    if label == "延期":
        return metrics.get("attention_level") == "overdue"
    if label == "多次延期":
        return bool(metrics.get("many_delays"))
    if label == "待性质标记":
        return _is_delay_nature_pending_from_data(merge_with_sample_order_template(record))
    if label == "平均考核分":
        score = metrics.get("assessment_score")
        return isinstance(score, int) and not isinstance(score, bool)
    return False


def get_record_revision(record: object) -> int:
    """读取记录版本号。"""
    if not isinstance(record, dict):
        return 0
    return max(0, normalize_int(record.get("_revision"), 0))


def _calculate_assessment_days_from_data(data: dict) -> Optional[int]:
    """使用已经标准化的样品单计算考核天数，避免重复深拷贝。"""
    basic = data["basic_info"]
    execution = data["execution"]
    actual = parse_iso_date(execution.get("actual_delivery_date"))
    planned = parse_iso_date(basic.get("planned_delivery_date"))
    if actual is None or planned is None:
        return None

    assessment_target = planned
    for extension in data["extensions"]:
        target_date = parse_iso_date(extension.get("target_date"))
        if target_date is not None and "主责" not in option_text(extension.get("reason")):
            assessment_target = target_date
    return (actual - assessment_target).days


def calculate_assessment_days(record: object) -> Optional[int]:
    """复刻原 Excel 的责任归属口径，计算实际交付相对考核基准的自然日差。"""
    return _calculate_assessment_days_from_data(merge_with_sample_order_template(record))


def calculate_assessment_score(days: Optional[int]) -> Optional[int]:
    """按原表考核区间把交付天数差换算为分数。"""
    if days is None:
        return None
    if days < -5:
        return 150
    if days < -3:
        return 140
    if days < 0:
        return 130
    if days == 0:
        return 120
    if days <= 3:
        return 100
    if days <= 5:
        return 80
    if days <= 10:
        return 60
    return 0


def calculate_sample_order_metrics(record: object, today: Optional[date] = None) -> dict:
    """计算一张样品单的全部派生字段与警示等级。"""
    data = merge_with_sample_order_template(record)
    basic = data["basic_info"]
    execution = data["execution"]
    extensions = data["extensions"]
    special_status = data["special_status"]["status"]
    today_value = today or date.today()

    application_date = parse_iso_date(basic.get("application_date"))
    planned_date = parse_iso_date(basic.get("planned_delivery_date"))
    actual_date = parse_iso_date(execution.get("actual_delivery_date"))
    extension_targets = [parse_iso_date(item.get("target_date")) for item in extensions]
    valid_extension_targets = [target for target in extension_targets if target is not None]
    warning_days = SAMPLE_ORDER_WARNING_DAYS

    cycle_days = None
    if application_date is not None and planned_date is not None:
        cycle_days = (planned_date - application_date).days

    target_date = valid_extension_targets[-1] if valid_extension_targets else planned_date
    delay_count = len(extensions)
    stage = f"第{delay_count}次延期" if delay_count else ""
    remaining_days = None
    if special_status == "正常" and actual_date is None and target_date is not None:
        # 历史目标只需要判定已经逾期，不必从历史日期逐日倒算到今天。
        remaining_days = -1 if target_date < today_value else business_days_between(today_value, target_date)

    if special_status == "作废":
        alert_message = "订单已作废"
        attention_level = "voided"
    elif special_status != "正常":
        alert_message = f"订单状态：{special_status}"
        attention_level = "paused"
    elif actual_date is not None:
        alert_message = "已完成"
        attention_level = "completed"
    elif planned_date is None:
        alert_message = "计划交货日期未填"
        attention_level = "missing"
    elif remaining_days is not None and remaining_days > warning_days:
        alert_message = f"{stage}充裕 剩{remaining_days}个工作日"
        attention_level = "normal"
    elif remaining_days is not None and remaining_days >= 0:
        alert_message = f"{stage}示警 剩{remaining_days}个工作日"
        attention_level = "warning"
    else:
        alert_message = f"第{delay_count + 1}次延期目标日期未填"
        attention_level = "overdue"

    assessment_days = _calculate_assessment_days_from_data(data)
    assessment_score = calculate_assessment_score(assessment_days)
    if special_status == "作废":
        assessment_days = None
        assessment_score = None
    if special_status != "正常":
        expected_status = special_status
    elif actual_date is not None:
        expected_status = "延期" if (assessment_days or 0) > 0 else "按期"
    elif target_date is None:
        expected_status = "待定"
    elif planned_date is not None and today_value <= planned_date:
        expected_status = "按计划"
    elif today_value <= target_date:
        expected_status = "按当前目标"
    else:
        expected_status = "延期"

    return {
        "cycle_days": cycle_days,
        "remaining_workdays": remaining_days,
        "alert_message": alert_message,
        "attention_level": attention_level,
        "expected_status": expected_status,
        "assessment_days": assessment_days,
        "assessment_score": assessment_score,
        "effective_target_date": target_date.isoformat() if target_date else "",
        "delay_count": delay_count,
        "many_delays": delay_count > SAMPLE_ORDER_DELAY_ATTENTION_THRESHOLD,
        "special_status": special_status,
    }


def sample_order_requires_attention(record: object, today: Optional[date] = None) -> bool:
    """判断记录是否需要在主页显示红色待办徽标。"""
    level = calculate_sample_order_metrics(record, today).get("attention_level")
    return level in {"missing", "warning", "overdue", "paused"}


def sample_order_is_overdue(record: object, today: Optional[date] = None) -> bool:
    """判断未完成订单是否已经超过当前有效目标日期。"""
    return calculate_sample_order_metrics(record, today).get("attention_level") == "overdue"


def _is_delay_nature_pending_from_data(data: dict) -> bool:
    """使用已经标准化的样品单判断是否待标记延期性质。"""
    actual_date = parse_iso_date(data["execution"].get("actual_delivery_date"))
    return bool(
        actual_date
        and data["extensions"]
        and data["special_status"].get("status") != "作废"
        and not option_text(data["delay_nature"].get("tag"))
    )


def is_delay_nature_pending(record: object) -> bool:
    """判断订单是否已交付、有延期且尚未由研发经理标记延期性质。"""
    return _is_delay_nature_pending_from_data(merge_with_sample_order_template(record))


def get_delay_nature_catalog(all_records: object) -> list[str]:
    """从历史订单标记中自动汇总性质标签，并按使用次数优先排序。"""
    if not isinstance(all_records, dict):
        return []
    tag_counts: dict[str, int] = {}
    for raw_record in all_records.values():
        if not isinstance(raw_record, dict) or not option_text(raw_record.get("record_id")):
            continue
        record = merge_with_sample_order_template(raw_record)
        nature = record["delay_nature"]
        tag = option_text(nature.get("tag"))
        if not tag:
            continue
        tag_counts[tag] = tag_counts.get(tag, 0) + 1
    return sorted(tag_counts, key=lambda tag: (-tag_counts[tag], tag))


def get_sample_order_dashboard_pending_count(
    all_records: object,
    today: Optional[date] = None,
    *,
    current_role: str = "",
) -> int:
    """按当前角色统计主页卡片需要显示的红点数量。"""
    if not isinstance(all_records, dict):
        return 0
    valid_records = [record for record in all_records.values() if isinstance(record, dict)]
    if is_sample_order_delay_nature_marker(current_role):
        return sum(1 for record in valid_records if is_delay_nature_pending(record))
    if not is_sample_order_delay_editor(current_role):
        return 0
    return sum(1 for record in valid_records if sample_order_is_overdue(record, today))


def validate_sample_order_submission(
    record: object,
    *,
    check_basic: bool,
    check_execution: bool,
    check_delay: bool,
    check_special_status: bool,
    today: Optional[date] = None,
) -> list[str]:
    """校验当前角色本次能够修改的字段组。"""
    data = merge_with_sample_order_template(record)
    basic = data["basic_info"]
    execution = data["execution"]
    extensions = data["extensions"]
    special_status = data["special_status"]
    errors: list[str] = []

    application_date = parse_iso_date(basic.get("application_date"))
    planned_date = parse_iso_date(basic.get("planned_delivery_date"))
    actual_date = parse_iso_date(execution.get("actual_delivery_date"))
    today_value = today or date.today()

    if check_basic:
        required = [
            ("样品单号", basic.get("sample_order_no")),
            ("客户编码", basic.get("customer_code")),
            ("产品型号", basic.get("product_model")),
            ("申请日期", basic.get("application_date")),
            ("申请人", basic.get("applicant")),
            ("计划交货日期", basic.get("planned_delivery_date")),
        ]
        missing = [label for label, value in required if not option_text(value)]
        if missing:
            errors.append(f"请填写：{'、'.join(missing)}")
        if normalize_int(basic.get("application_qty"), 0) <= 0:
            errors.append("申请数量必须大于 0")
        if application_date is None and option_text(basic.get("application_date")):
            errors.append("申请日期格式不正确")
        if planned_date is None and option_text(basic.get("planned_delivery_date")):
            errors.append("计划交货日期格式不正确")
        if application_date and planned_date and planned_date < application_date:
            errors.append("计划交货日期不能早于申请日期")

    if check_execution:
        if option_text(execution.get("actual_delivery_date")) and actual_date is None:
            errors.append("实际交货日期格式不正确")
        if actual_date and application_date and actual_date < application_date:
            errors.append("实际交货日期不能早于申请日期")
        if actual_date and actual_date > today_value:
            errors.append("实际交货日期不能晚于当天")

    if check_delay:
        for index, extension in enumerate(extensions, start=1):
            target_text = option_text(extension.get("target_date"))
            reason = option_text(extension.get("reason"))
            target_date = parse_iso_date(target_text)
            if not target_text or not reason:
                errors.append(f"第{index}次延期的目标日期和原因必须完整填写")
                continue
            if target_date is None:
                errors.append(f"第{index}次延期目标日期格式不正确")
                continue
            if not option_text(extension.get("extension_id")) and target_date < today_value:
                errors.append(f"第{index}次延期目标日期不能早于当天")

    if check_special_status:
        status = option_text(special_status.get("status"), "正常")
        reason = option_text(special_status.get("reason"))
        if status not in SAMPLE_ORDER_SPECIAL_STATUSES:
            errors.append("订单特殊状态不在后台配置的允许范围内")
        if SAMPLE_ORDER_SPECIAL_STATUS_REASON_REQUIRED and status != "正常" and not reason:
            errors.append("设置订单特殊状态时必须填写原因")
    return errors


def get_sample_order_dashboard_url(record_id: str = "") -> str:
    """生成企业微信消息中的样品单直达链接。"""
    page_url = f"{SAMPLE_ORDER_PUBLIC_BASE_URL}/sample_order_dashboard"
    return f"{page_url}?record_id={quote(record_id, safe='')}" if record_id else page_url


async def _send_sample_order_change_notifications(
    record: dict,
    extension_events: list[dict],
    status_event: Optional[dict],
) -> tuple[str, ...]:
    """在数据提交成功后发送延期和特殊状态企业微信通知。"""
    basic = record["basic_info"]
    record_id = option_text(record.get("record_id"))
    applicant = option_text(basic.get("applicant"))
    applicant_needed = bool(extension_events and SAMPLE_ORDER_NOTIFY_APPLICANT_ON_EXTENSION) or bool(
        status_event and SAMPLE_ORDER_NOTIFY_APPLICANT_ON_SPECIAL_STATUS
    )
    redirect_applicant = bool(applicant_needed and SAMPLE_ORDER_REDIRECT_APPLICANT_NOTIFICATIONS_TO_MANAGER)
    applicant_recipient = ""
    if applicant and applicant_needed and not redirect_applicant:
        applicant_recipient = await resolve_wecom_recipients(
            [{"names": [applicant]}],
            fallback_touser="",
        )
    manager_recipients = ""
    needs_manager = (
        redirect_applicant
        or bool(status_event)
        or any(
            normalize_int(event.get("extension_number"), 0) > SAMPLE_ORDER_DELAY_ATTENTION_THRESHOLD
            for event in extension_events
        )
    )
    if needs_manager:
        manager_recipients = await resolve_wecom_recipients(
            SAMPLE_ORDER_MANAGER_NOTIFY_TARGETS,
            fallback_touser="",
        )

    failures: list[str] = []
    if applicant_needed and not redirect_applicant and not applicant_recipient:
        failures.append(f"申请人“{applicant or '未填写'}”未匹配到企业微信成员")
    if needs_manager and not manager_recipients:
        failures.append("研发经理通知规则未匹配到企业微信成员")
    for event in extension_events:
        extension_number = normalize_int(event.get("extension_number"), 0)
        notify_manager = extension_number > SAMPLE_ORDER_DELAY_ATTENTION_THRESHOLD
        recipients = merge_wecom_recipients(
            applicant_recipient if SAMPLE_ORDER_NOTIFY_APPLICANT_ON_EXTENSION and not redirect_applicant else "",
            manager_recipients if notify_manager or redirect_applicant else "",
        )
        if not recipients:
            failures.append(f"第{extension_number}次延期未匹配到企业微信收件人")
            continue
        manager_notice = (
            f"\n该订单已延期{extension_number}次，超过配置阈值"
            f"{SAMPLE_ORDER_DELAY_ATTENTION_THRESHOLD}次，请研发经理关注。"
            if notify_manager
            else ""
        )
        content = (
            "【样品单延期通知】\n"
            f"样品单号：{basic.get('sample_order_no', '')}\n"
            f"产品型号：{basic.get('product_model', '')}\n"
            f"申请人：{applicant}\n"
            f"延期次数：第{extension_number}次\n"
            f"延期目标日期：{event.get('target_date', '')}\n"
            f"延期原因：{event.get('reason', '')}\n"
            f"操作人：{event.get('created_by', '')}（{event.get('created_role', '')}）"
            f"{manager_notice}"
        )
        success, message = await send_wecom_text_message(
            content,
            recipients,
            module="sample_order_dashboard",
            business_key=f"{record_id}:extension:{event.get('extension_id', '')}",
            message_type="extension_created",
            link_url=get_sample_order_dashboard_url(record_id),
        )
        if not success:
            failures.append(f"第{extension_number}次延期通知失败：{message}")

    if status_event:
        recipients = merge_wecom_recipients(
            applicant_recipient if SAMPLE_ORDER_NOTIFY_APPLICANT_ON_SPECIAL_STATUS and not redirect_applicant else "",
            manager_recipients,
        )
        if not recipients:
            failures.append("特殊状态通知未匹配到企业微信收件人")
        else:
            content = (
                "【样品单特殊状态通知】\n"
                f"样品单号：{basic.get('sample_order_no', '')}\n"
                f"产品型号：{basic.get('product_model', '')}\n"
                f"申请人：{applicant}\n"
                f"状态变更：{status_event.get('old_status', '')} → {status_event.get('status', '')}\n"
                f"设置原因：{status_event.get('reason', '') or '无'}\n"
                f"操作人：{status_event.get('updated_by', '')}（{status_event.get('updated_role', '')}）"
            )
            success, message = await send_wecom_text_message(
                content,
                recipients,
                module="sample_order_dashboard",
                business_key=f"{record_id}:special_status:{status_event.get('history_id', '')}",
                message_type="special_status_changed",
                link_url=get_sample_order_dashboard_url(record_id),
            )
            if not success:
                failures.append(f"特殊状态通知失败：{message}")
    return tuple(failures)


async def _send_sample_order_notifications_in_background(
    record: dict,
    extension_events: list[dict],
    status_event: Optional[dict],
) -> None:
    """后台发送样品单通知；失败由统一重试记录接管并写入日志。"""
    failures = await _send_sample_order_change_notifications(
        record,
        extension_events,
        status_event,
    )
    if failures:
        logger.warning("样品单后台通知未即时成功：%s", "；".join(failures))


async def save_sample_order_record(
    submitted: dict,
    user: str,
    role: str,
    *,
    is_new: bool,
) -> SampleOrderUpdateResult:
    """按字段职责和记录版本原子保存一张样品单。"""
    can_edit_base = is_sample_order_base_editor(role)
    can_edit_delay = is_sample_order_delay_editor(role)
    can_edit_special_status = is_sample_order_special_status_editor(role)
    can_edit_execution = can_edit_base
    if is_new and not can_edit_base:
        return SampleOrderUpdateResult(True, False, "forbidden")
    if not (can_edit_base or can_edit_delay or can_edit_special_status):
        return SampleOrderUpdateResult(True, False, "forbidden")

    record = merge_with_sample_order_template(submitted)
    errors = validate_sample_order_submission(
        record,
        check_basic=can_edit_base,
        check_execution=can_edit_execution,
        check_delay=can_edit_delay,
        check_special_status=can_edit_special_status,
    )
    if errors:
        return SampleOrderUpdateResult(True, False, "invalid", {"errors": errors})

    record_id = option_text(record.get("record_id")) or uuid.uuid4().hex
    expected_revision = None if is_new else get_record_revision(record)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    outcome: dict[str, Any] = {
        "changed": False,
        "code": "db_error",
        "record": None,
        "extension_events": [],
        "status_event": None,
    }

    def apply_update(current: object) -> object:
        current_exists = isinstance(current, dict) and bool(option_text(current.get("record_id")))
        if is_new:
            if current is not None:
                outcome["code"] = "already_exists"
                return db_storage.ATOMIC_NO_UPDATE
            updated = get_sample_order_template()
            updated["record_id"] = record_id
            updated["created_by"] = user
            updated["created_role"] = role
            updated["created_at"] = now_str
        else:
            if not current_exists:
                outcome["code"] = "not_found"
                return db_storage.ATOMIC_NO_UPDATE
            updated = merge_with_sample_order_template(current)
            if get_record_revision(updated) != expected_revision:
                outcome["code"] = "revision_conflict"
                outcome["record"] = copy.deepcopy(updated)
                return db_storage.ATOMIC_NO_UPDATE

        changed_sections: list[str] = []
        if can_edit_base:
            updated["basic_info"] = copy.deepcopy(record["basic_info"])
            changed_sections.append("基础信息")
        if can_edit_execution:
            updated["execution"] = copy.deepcopy(record["execution"])
            changed_sections.append("执行信息")
        if can_edit_delay:
            stored_extensions = [normalize_extension(item) for item in updated.get("extensions", [])]
            incoming_extensions = [normalize_extension(item) for item in record.get("extensions", [])]
            if len(incoming_extensions) < len(stored_extensions):
                outcome["code"] = "extension_history_conflict"
                outcome["record"] = copy.deepcopy(updated)
                return db_storage.ATOMIC_NO_UPDATE
            if incoming_extensions[: len(stored_extensions)] != stored_extensions:
                outcome["code"] = "extension_history_conflict"
                outcome["record"] = copy.deepcopy(updated)
                return db_storage.ATOMIC_NO_UPDATE
            new_extensions = incoming_extensions[len(stored_extensions) :]
            for extension in new_extensions:
                if extension["extension_id"]:
                    outcome["code"] = "extension_history_conflict"
                    outcome["record"] = copy.deepcopy(updated)
                    return db_storage.ATOMIC_NO_UPDATE
                extension["extension_id"] = uuid.uuid4().hex
                extension["created_by"] = user
                extension["created_role"] = role
                extension["created_at"] = now_str
                extension_event = copy.deepcopy(extension)
                extension_event["extension_number"] = len(stored_extensions) + 1
                stored_extensions.append(extension)
                outcome["extension_events"].append(extension_event)
            updated["extensions"] = stored_extensions
            if new_extensions:
                changed_sections.append("延期信息")
        if can_edit_special_status:
            stored_special = updated["special_status"]
            incoming_special = record["special_status"]
            old_status = option_text(stored_special.get("status"), "正常")
            new_status = option_text_in(incoming_special.get("status"), SAMPLE_ORDER_SPECIAL_STATUSES, "正常")
            new_reason = option_text(incoming_special.get("reason"))
            status_changed = new_status != old_status or new_reason != option_text(stored_special.get("reason"))
            if status_changed:
                history_item = {
                    "history_id": uuid.uuid4().hex,
                    "old_status": old_status,
                    "status": new_status,
                    "reason": new_reason,
                    "updated_by": user,
                    "updated_role": role,
                    "updated_at": now_str,
                }
                history = stored_special.get("history", [])
                stored_special["history"] = copy.deepcopy(history) if isinstance(history, list) else []
                stored_special["history"].append(copy.deepcopy(history_item))
                stored_special.update(
                    {
                        "status": new_status,
                        "reason": new_reason,
                        "updated_by": user,
                        "updated_role": role,
                        "updated_at": now_str,
                    }
                )
                outcome["status_event"] = copy.deepcopy(history_item)
                changed_sections.append("特殊状态")

        updated["updated_by"] = user
        updated["updated_role"] = role
        updated["updated_at"] = now_str
        updated["_revision"] = get_record_revision(updated) + 1
        updated.setdefault("operation_log", []).append(
            {
                "user": user,
                "role": role,
                "action": f"保存{'、'.join(changed_sections) or '记录'}",
                "time": now_str,
            }
        )
        outcome["changed"] = True
        outcome["code"] = "created" if is_new else "updated"
        outcome["record"] = copy.deepcopy(updated)
        return updated

    success = await db_storage.atomic_json_entity_update(
        SAMPLE_ORDER_ENTITY_NAMESPACE,
        record_id,
        apply_update,
    )
    if success and outcome["changed"]:
        await db_storage.set_item(SAMPLE_ORDER_VERSION_KEY, time.time())
        if outcome["record"] and (outcome["extension_events"] or outcome["status_event"]):
            schedule_background_task(
                _send_sample_order_notifications_in_background(
                    copy.deepcopy(outcome["record"]),
                    copy.deepcopy(outcome["extension_events"]),
                    copy.deepcopy(outcome["status_event"]),
                ),
                f"样品单{record_id}企业微信通知",
            )
    return SampleOrderUpdateResult(
        db_success=success,
        changed=bool(success and outcome["changed"]),
        code=outcome["code"] if success else "db_error",
        record=outcome["record"],
    )


async def import_sample_order_records(
    records: list[dict],
    user: str,
    role: str,
    *,
    source_name: str,
) -> SampleOrderImportResult:
    """由研发助理把预览通过的Excel记录一次性原子导入。"""
    if not is_sample_order_base_editor(role):
        return SampleOrderImportResult(True, 0, "forbidden")
    if not records:
        return SampleOrderImportResult(True, 0, "empty")

    normalized_records = [merge_with_sample_order_template(record) for record in records]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    existing_ids = set(get_all_sample_order_records())
    imported_entities: dict[str, Any] = {}
    for source_record in normalized_records:
        imported = merge_with_sample_order_template(source_record)
        record_id = uuid.uuid4().hex
        while record_id in existing_ids or record_id in imported_entities:
            record_id = uuid.uuid4().hex
        imported["record_id"] = record_id
        imported["created_by"] = user
        imported["created_role"] = role
        imported["created_at"] = now_str
        imported["updated_by"] = user
        imported["updated_role"] = role
        imported["updated_at"] = now_str
        imported["_revision"] = 1
        imported["execution"]["sample_owner"] = SAMPLE_ORDER_EXCEL_IMPORT_OWNER
        imported["import_info"]["source_name"] = source_name
        imported_extensions: list[dict] = []
        for extension in imported["extensions"]:
            normalized_extension = normalize_extension(extension)
            normalized_extension.update(
                {
                    "extension_id": uuid.uuid4().hex,
                    "created_by": user,
                    "created_role": role,
                    "created_at": now_str,
                }
            )
            imported_extensions.append(normalized_extension)
        imported["extensions"] = imported_extensions
        imported["operation_log"] = [
            {
                "user": user,
                "role": role,
                "action": f"从Excel导入（第{normalize_int(imported['import_info'].get('source_row'), 0)}行）",
                "time": now_str,
            }
        ]
        imported_entities[record_id] = imported

    success = await db_storage.insert_json_entities(
        SAMPLE_ORDER_ENTITY_NAMESPACE,
        imported_entities,
    )
    if success:
        await db_storage.set_item(SAMPLE_ORDER_VERSION_KEY, time.time())
    return SampleOrderImportResult(
        db_success=success,
        imported_count=len(imported_entities) if success else 0,
        code="imported" if success else "db_error",
    )


async def mark_sample_order_delay_nature(
    record_id: str,
    nature_tag: str,
    user: str,
    role: str,
    *,
    expected_revision: int,
) -> SampleOrderUpdateResult:
    """由研发经理原子标记已完成延期订单的标准性质标签。"""
    if not is_sample_order_delay_nature_marker(role):
        return SampleOrderUpdateResult(True, False, "forbidden")
    normalized_tag = option_text(nature_tag)
    if not normalized_tag:
        return SampleOrderUpdateResult(True, False, "invalid", {"errors": ["延期性质标签不能为空"]})
    if len(normalized_tag) > 50:
        return SampleOrderUpdateResult(True, False, "invalid", {"errors": ["延期性质标签不能超过50个字符"]})

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    outcome: dict[str, Any] = {"changed": False, "code": "db_error", "record": None}

    def apply_mark(current: object) -> object:
        if not isinstance(current, dict) or not option_text(current.get("record_id")):
            outcome["code"] = "not_found"
            return db_storage.ATOMIC_NO_UPDATE
        record = merge_with_sample_order_template(current)
        if get_record_revision(record) != expected_revision:
            outcome["code"] = "revision_conflict"
            outcome["record"] = copy.deepcopy(record)
            return db_storage.ATOMIC_NO_UPDATE
        if (
            parse_iso_date(record["execution"].get("actual_delivery_date")) is None
            or not record["extensions"]
            or record["special_status"].get("status") == "作废"
        ):
            outcome["code"] = "nature_not_applicable"
            outcome["record"] = copy.deepcopy(record)
            return db_storage.ATOMIC_NO_UPDATE

        nature = record["delay_nature"]
        old_tag = option_text(nature.get("tag"))
        if old_tag == normalized_tag:
            outcome["code"] = "no_change"
            outcome["record"] = copy.deepcopy(record)
            return db_storage.ATOMIC_NO_UPDATE
        history_item = {
            "history_id": uuid.uuid4().hex,
            "old_tag": old_tag,
            "tag": normalized_tag,
            "marked_by": user,
            "marked_role": role,
            "marked_at": now_str,
        }
        history = nature.get("history", [])
        nature["history"] = copy.deepcopy(history) if isinstance(history, list) else []
        nature["history"].append(copy.deepcopy(history_item))
        nature.update(
            {
                "tag": normalized_tag,
                "marked_by": user,
                "marked_role": role,
                "marked_at": now_str,
            }
        )
        record["updated_by"] = user
        record["updated_role"] = role
        record["updated_at"] = now_str
        record["_revision"] = get_record_revision(record) + 1
        record.setdefault("operation_log", []).append(
            {
                "user": user,
                "role": role,
                "action": f"标记延期性质：{normalized_tag}",
                "time": now_str,
            }
        )
        outcome["changed"] = True
        outcome["code"] = "updated"
        outcome["record"] = copy.deepcopy(record)
        return record

    success = await db_storage.atomic_json_entity_update(
        SAMPLE_ORDER_ENTITY_NAMESPACE,
        record_id,
        apply_mark,
    )
    if success and outcome["changed"]:
        await db_storage.set_item(SAMPLE_ORDER_VERSION_KEY, time.time())
    return SampleOrderUpdateResult(
        db_success=success,
        changed=bool(success and outcome["changed"]),
        code=outcome["code"] if success else "db_error",
        record=outcome["record"],
    )


async def delete_sample_order_record(record_id: str, role: str) -> SampleOrderUpdateResult:
    """由管理员原子删除一张样品单。"""
    if not is_sample_order_admin(role):
        return SampleOrderUpdateResult(True, False, "forbidden")
    outcome: dict[str, Any] = {"changed": False, "code": "db_error", "record": None}

    def remove_record(current: object) -> object:
        if not isinstance(current, dict):
            outcome["code"] = "not_found"
            return db_storage.ATOMIC_NO_UPDATE
        outcome["record"] = copy.deepcopy(current)
        outcome["changed"] = True
        outcome["code"] = "deleted"
        return db_storage.ATOMIC_DELETE

    success = await db_storage.atomic_json_entity_update(
        SAMPLE_ORDER_ENTITY_NAMESPACE,
        record_id,
        remove_record,
    )
    if success and outcome["changed"]:
        await db_storage.set_item(SAMPLE_ORDER_VERSION_KEY, time.time())
    return SampleOrderUpdateResult(
        db_success=success,
        changed=bool(success and outcome["changed"]),
        code=outcome["code"] if success else "db_error",
        record=outcome["record"],
    )


def build_sample_order_row(
    record: object,
    today: Optional[date] = None,
    *,
    calculated_metrics: Optional[dict] = None,
) -> dict:
    """把存储记录整理成表格行数据。"""
    data = merge_with_sample_order_template(record)
    basic = data["basic_info"]
    execution = data["execution"]
    extensions = data["extensions"]
    special_status = data["special_status"]
    delay_nature = data["delay_nature"]
    metrics = (
        calculated_metrics if isinstance(calculated_metrics, dict) else calculate_sample_order_metrics(data, today)
    )
    assessment_days = metrics["assessment_days"]
    assessment_score = metrics["assessment_score"]
    return {
        "record_id": data["record_id"],
        "sample_order_no": basic["sample_order_no"],
        "customer_code": basic["customer_code"],
        "product_model": basic["product_model"],
        "application_qty": basic["application_qty"],
        "application_date": basic["application_date"],
        "applicant": basic["applicant"],
        "planned_delivery_date": basic["planned_delivery_date"],
        "cycle_days": metrics["cycle_days"] if metrics["cycle_days"] is not None else "",
        "remark": basic["remark"],
        "actual_delivery_date": execution["actual_delivery_date"],
        "warning_days": SAMPLE_ORDER_WARNING_DAYS,
        "alert_message": metrics["alert_message"],
        "sample_owner": execution["sample_owner"],
        "expected_status": metrics["expected_status"],
        "delay_count": len(extensions),
        "latest_extension_target": extensions[-1]["target_date"] if extensions else "",
        "latest_extension_reason": extensions[-1]["reason"] if extensions else "",
        "special_status": special_status["status"],
        "special_status_reason": special_status["reason"],
        "delay_nature_tag": delay_nature["tag"],
        "nature_pending": _is_delay_nature_pending_from_data(data),
        "assessment_days": assessment_days if assessment_days is not None else "",
        "assessment_score": assessment_score if assessment_score is not None else "",
        "attention_level": metrics["attention_level"],
        "updated_at": data["updated_at"],
    }


def get_sample_order_card_palette(
    attention_level: object,
    *,
    nature_pending: bool,
    can_mark_delay_nature: bool,
) -> tuple[str, str]:
    """按订单状态和当前角色返回卡片背景类及左边框颜色。"""
    level = option_text(attention_level, "normal")
    if nature_pending and can_mark_delay_nature:
        return (
            "bg-yellow-50 border-yellow-300 shadow-md hover:bg-yellow-100",
            "#eab308",
        )
    palette = {
        "overdue": (
            "bg-red-50 border-red-300 shadow-md hover:bg-red-100",
            "#ef4444",
        ),
        "missing": (
            "bg-red-50 border-red-300 shadow-md hover:bg-red-100",
            "#ef4444",
        ),
        "warning": (
            "bg-orange-50 border-orange-300 shadow-md hover:bg-orange-100",
            "#f97316",
        ),
        "paused": (
            "bg-purple-50 border-purple-300 shadow-sm hover:bg-purple-100",
            "#a855f7",
        ),
        "completed": (
            "bg-green-50 border-green-300 shadow-sm hover:bg-green-100",
            "#22c55e",
        ),
        "voided": (
            "bg-gray-100 border-gray-300 shadow-sm hover:bg-gray-200",
            "#9ca3af",
        ),
    }
    return palette.get(
        level,
        ("bg-white border-gray-200 shadow-sm hover:bg-blue-50", "#3b82f6"),
    )


def sample_order_matches_filter(
    record: object,
    filter_value: str,
    today: Optional[date] = None,
    *,
    calculated_metrics: Optional[dict] = None,
) -> bool:
    """判断记录是否符合页面状态筛选条件。"""
    data = merge_with_sample_order_template(record)
    actual_date = parse_iso_date(data["execution"].get("actual_delivery_date"))
    if filter_value == FILTER_COMPLETED:
        return actual_date is not None
    if filter_value == FILTER_IN_PROGRESS:
        return bool(actual_date is None and data["special_status"].get("status") == "正常")
    if filter_value == FILTER_NATURE_PENDING:
        return _is_delay_nature_pending_from_data(data)
    if filter_value in SAMPLE_ORDER_SPECIAL_STATUSES and filter_value != "正常":
        return data["special_status"].get("status") == filter_value
    if filter_value == FILTER_ALL:
        return True

    metrics = (
        calculated_metrics if isinstance(calculated_metrics, dict) else calculate_sample_order_metrics(data, today)
    )
    level = metrics["attention_level"]
    if filter_value == FILTER_WARNING:
        return level in {"missing", "warning"}
    if filter_value == FILTER_DELAYED:
        return level == "overdue" or metrics["expected_status"] == "延期"
    if filter_value == FILTER_MANY_DELAYS:
        return bool(metrics["many_delays"])
    return True


def get_sample_order_monthly_statistics(
    all_records: object,
    today: Optional[date] = None,
    *,
    date_basis: str = "planned",
    count_basis: str = "orders",
) -> list[dict[str, object]]:
    """按计划或实际交货月份，统计订单数或申请样品数。"""
    normalized_basis = "actual" if date_basis == "actual" else "planned"
    normalized_count_basis = "samples" if count_basis == "samples" else "orders"
    reference_date = today or date.today()
    reference_month_number = reference_date.year * 12 + reference_date.month - 1

    if isinstance(all_records, dict):
        raw_records = list(all_records.values())
    elif isinstance(all_records, (list, tuple)):
        raw_records = list(all_records)
    else:
        raw_records = []

    normalized_records: list[tuple[dict, date, Optional[date]]] = []
    latest_month_number = reference_month_number
    for raw_record in raw_records:
        if not isinstance(raw_record, dict):
            continue
        record = merge_with_sample_order_template(raw_record)
        planned_date = parse_iso_date(record["basic_info"].get("planned_delivery_date"))
        if planned_date is None:
            continue
        actual_date = parse_iso_date(record["execution"].get("actual_delivery_date"))
        grouping_date = planned_date if normalized_basis == "planned" else actual_date
        if grouping_date is None:
            # 未完成订单没有实际交货日期，无法归入实际交样月份。
            continue
        normalized_records.append((record, planned_date, actual_date))
        grouping_month_number = grouping_date.year * 12 + grouping_date.month - 1
        latest_month_number = max(latest_month_number, grouping_month_number)

    statistics = [
        {
            "month": f"{month_number // 12:04d}-{month_number % 12 + 1:02d}",
            "on_time_completed": 0,
            "delayed_completed": 0,
            "incomplete": 0,
            "total": 0,
        }
        for month_number in range(reference_month_number - 11, latest_month_number + 1)
    ]
    statistics_by_month = {item["month"]: item for item in statistics}

    for record, planned_date, actual_date in normalized_records:
        grouping_date = planned_date if normalized_basis == "planned" else actual_date
        if grouping_date is None:
            continue
        month_key = grouping_date.strftime("%Y-%m")
        month_statistics = statistics_by_month.get(month_key)
        if month_statistics is None:
            continue

        if actual_date is None:
            category = "incomplete"
        elif actual_date <= planned_date:
            category = "on_time_completed"
        else:
            category = "delayed_completed"
        increment = (
            max(0, normalize_int(record["basic_info"].get("application_qty"), 0))
            if normalized_count_basis == "samples"
            else 1
        )
        month_statistics[category] = normalize_int(month_statistics[category], 0) + increment
        month_statistics["total"] = normalize_int(month_statistics["total"], 0) + increment
    return statistics


def get_sample_order_statistics_details(
    all_records: object,
    month: str,
    category: str,
    *,
    date_basis: str = "planned",
) -> list[dict[str, object]]:
    """返回统计图指定月份和完成分类所包含的订单明细。"""
    normalized_basis = "actual" if date_basis == "actual" else "planned"
    category_keys = {
        "按时完成": "on_time_completed",
        "延期完成": "delayed_completed",
        "未完成": "incomplete",
    }
    target_category = category_keys.get(category)
    if target_category is None:
        return []

    if isinstance(all_records, dict):
        raw_records = all_records.values()
    elif isinstance(all_records, (list, tuple)):
        raw_records = all_records
    else:
        raw_records = []

    details: list[dict[str, object]] = []
    for raw_record in raw_records:
        if not isinstance(raw_record, dict):
            continue
        record = merge_with_sample_order_template(raw_record)
        basic = record["basic_info"]
        execution = record["execution"]
        planned_date = parse_iso_date(basic.get("planned_delivery_date"))
        actual_date = parse_iso_date(execution.get("actual_delivery_date"))
        if planned_date is None:
            continue
        grouping_date = planned_date if normalized_basis == "planned" else actual_date
        if grouping_date is None or grouping_date.strftime("%Y-%m") != month:
            continue

        if actual_date is None:
            record_category = "incomplete"
        elif actual_date <= planned_date:
            record_category = "on_time_completed"
        else:
            record_category = "delayed_completed"
        if record_category != target_category:
            continue
        details.append(
            {
                "record_id": option_text(record.get("record_id")),
                "sample_order_no": option_text(basic.get("sample_order_no"), "-"),
                "product_model": option_text(basic.get("product_model"), "-"),
                "application_qty": max(0, normalize_int(basic.get("application_qty"), 0)),
                "applicant": option_text(basic.get("applicant"), "-"),
            }
        )
    return sorted(details, key=lambda item: option_text(item.get("sample_order_no")))


def build_sample_order_statistics_chart(
    statistics: list[dict[str, object]],
    *,
    include_incomplete: bool = True,
    value_name: str = "按订单数",
) -> dict:
    """生成近12个月订单统计的堆叠柱状图配置。"""
    months = [option_text(item.get("month")) for item in statistics]
    series_meta = [
        ("按时完成", "on_time_completed", "#22c55e"),
        ("延期完成", "delayed_completed", "#f97316"),
    ]
    if include_incomplete:
        series_meta.append(("未完成", "incomplete", "#64748b"))
    series = [
        {
            "name": label,
            "type": "bar",
            "stack": "statistics",
            "barWidth": "52%",
            "data": [normalize_int(item.get(key), 0) for item in statistics],
            "itemStyle": {"color": color},
            "emphasis": {"focus": "series"},
        }
        for label, key, color in series_meta
    ]
    # 用透明散点把总数标签稳定地放在每根堆叠柱顶部，即使最上层分类数量为0也能正确显示。
    series.append(
        {
            "name": f"总{value_name}",
            "type": "scatter",
            "data": [normalize_int(item.get("total"), 0) for item in statistics],
            "symbolSize": 1,
            "silent": True,
            "itemStyle": {"color": "transparent"},
            "label": {
                "show": True,
                "position": "top",
                "distance": 6,
                "formatter": "{c}",
                "color": "#334155",
                "fontWeight": "bold",
            },
            "tooltip": {"show": False},
            "z": 10,
        }
    )
    return {
        "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}},
        "legend": {"top": 0, "data": [item[0] for item in series_meta]},
        "grid": {"top": 50, "bottom": 35, "left": 30, "right": 25, "containLabel": True},
        "xAxis": {
            "type": "category",
            "data": months,
            "axisTick": {"show": False},
            "axisLabel": {"interval": 0, "rotate": 30},
        },
        "yAxis": {
            "type": "value",
            "name": value_name,
            "minInterval": 1,
            "splitLine": {"lineStyle": {"type": "dashed"}},
        },
        "series": series,
    }


@ui.page("/sample_order_dashboard")
async def sample_order_dashboard_page(record_id: str = "") -> None:
    """构建样品单执行看板页面。"""
    setup_global_activity_tracking()
    ui.add_head_html(
        """
        <style>
            html, body { overflow: hidden !important; }
            .sample-order-card:hover { transform: translateY(-1px); }
        </style>
        """
    )
    if not app.storage.user.get("current_user"):
        redirect_target = f"/sample_order_dashboard?record_id={record_id}" if record_id else "/sample_order_dashboard"
        ui.navigate.to(f"/login?redirect_to={quote(redirect_target, safe='')}")
        return

    current_user = option_text(app.storage.user.get("current_user"), "未知用户")
    # 用户会话会跨服务器重启保留；每次进入权限页都从当前用户表同步角色，避免管理员改完角色后
    # 浏览器仍沿用旧的 current_role。
    current_role = option_text(sync_current_user_role(), "未知角色")
    current_display_path = get_cache_busted_path(
        app.storage.general.get("user_preferences", {}).get(current_user, {}).get("avatar", PRESET_AVATARS[0])
    )
    can_edit_base = is_sample_order_base_editor(current_role)
    can_edit_delay = is_sample_order_delay_editor(current_role)
    can_edit_special_status = is_sample_order_special_status_editor(current_role)
    can_mark_delay_nature = is_sample_order_delay_nature_marker(current_role)
    can_delete = is_sample_order_admin(current_role)
    page_state: dict[str, object] = {
        "search_keyword": "",
        "filter_state": DEFAULT_SAMPLE_ORDER_FILTER,
        "page": 1,
        "last_version": db_storage.get_item(SAMPLE_ORDER_VERSION_KEY, 0),
        "kpi_cache_key": None,
    }
    dashboard_data_cache: dict[str, Any] = {
        "key": None,
        "records_with_metrics": [],
    }
    detail_dialog = ui.dialog().props("maximized persistent")
    confirm_dialog = ui.dialog().props("persistent")
    import_dialog = ui.dialog().props("persistent")
    statistics_dialog = ui.dialog()
    statistics_detail_dialog = ui.dialog()

    def open_statistics_dialog() -> None:
        """打开可切换计划/实际交样月份口径的订单统计。"""
        all_records = get_all_sample_order_records()
        statistics_dialog.clear()
        with statistics_dialog, ui.card().classes("w-[1100px] max-w-[96vw] h-[680px] max-h-[92vh] p-5"):
            with ui.row().classes("w-full items-center justify-between shrink-0"):
                with ui.row().classes("items-center gap-2"):
                    ui.icon("stacked_bar_chart", color="blue", size="md")
                    ui.label("样品订单月度统计").classes("text-xl font-bold")
                with ui.row().classes("items-center gap-2"):
                    date_basis_toggle = (
                        ui.toggle(
                            {"planned": "按计划交样日期", "actual": "按实际交样日期"},
                            value="planned",
                        )
                        .props("color=grey-3 text-color=grey-5 toggle-color=teal toggle-text-color=white unelevated")
                        .classes("self-center shrink-0 text-sm")
                    )
                    count_basis_toggle = (
                        ui.toggle(
                            {"orders": "按订单数", "samples": "按样品数"},
                            value="orders",
                        )
                        .props("color=grey-3 text-color=grey-5 toggle-color=teal toggle-text-color=white unelevated")
                        .classes("self-center shrink-0 text-sm")
                    )
                ui.button(icon="close", on_click=statistics_dialog.close).props("flat round")
            chart_container = ui.column().classes("w-full flex-grow min-h-0 gap-1")

            def open_statistics_details(
                month: str,
                category: str,
                date_basis: str,
                count_basis: str,
            ) -> None:
                details = get_sample_order_statistics_details(
                    all_records,
                    month,
                    category,
                    date_basis=date_basis,
                )
                sample_total = sum(normalize_int(item.get("application_qty"), 0) for item in details)
                basis_label = "按实际交样日期" if date_basis == "actual" else "按计划交样日期"
                summary = (
                    f"共 {len(details)} 张订单、{sample_total} 个样品"
                    if count_basis == "samples"
                    else f"共 {len(details)} 张订单"
                )
                statistics_detail_dialog.clear()
                with statistics_detail_dialog, ui.card().classes("w-[900px] max-w-[95vw] max-h-[85vh] p-5"):
                    with ui.row().classes("w-full items-center justify-between mb-2"):
                        with ui.column().classes("gap-0"):
                            ui.label(f"{month} · {category}").classes("text-xl font-bold")
                            ui.label(f"{basis_label} · {summary}").classes("text-sm text-gray-500")
                        ui.button(icon="close", on_click=statistics_detail_dialog.close).props("flat round")
                    with ui.element("div").classes(
                        "grid grid-cols-[minmax(150px,1fr)_minmax(180px,1.5fr)_100px_minmax(120px,1fr)] "
                        "w-full gap-3 px-3 py-2 bg-slate-100 rounded-t font-bold text-sm text-slate-700"
                    ):
                        ui.label("订单号")
                        ui.label("产品型号")
                        ui.label("数量")
                        ui.label("申请人")
                    with ui.scroll_area().classes("w-full h-[55vh] border rounded-b"):
                        if not details:
                            ui.label("当前分类暂无订单").classes("w-full text-center text-gray-400 py-8")
                        for item in details:
                            with ui.element("div").classes(
                                "grid grid-cols-[minmax(150px,1fr)_minmax(180px,1.5fr)_100px_minmax(120px,1fr)] "
                                "w-full gap-3 px-3 py-2 border-b text-sm items-center hover:bg-blue-50"
                            ):
                                ui.label(option_text(item.get("sample_order_no"), "-")).classes("font-mono")
                                ui.label(option_text(item.get("product_model"), "-")).classes("break-all")
                                ui.label(str(normalize_int(item.get("application_qty"), 0)))
                                ui.label(option_text(item.get("applicant"), "-"))
                statistics_detail_dialog.open()

            def bind_statistics_chart_click(chart: Any, callback: Any) -> None:
                """把 ECharts 堆叠柱点击事件转发给 NiceGUI。"""
                chart.on("sample_order_statistics_click", callback)
                ui.run_javascript(f"""
                    setTimeout(() => {{
                        const el = getElement({chart.id});
                        if (!el || !el.chart) return;
                        if (el.__sampleOrderStatisticsHandler) {{
                            el.chart.off('click', el.__sampleOrderStatisticsHandler);
                        }}
                        el.__sampleOrderStatisticsHandler = function(params) {{
                            if (params.componentType === 'series' && params.seriesType === 'bar') {{
                                el.$emit('sample_order_statistics_click', {{
                                    month: params.name,
                                    category: params.seriesName
                                }});
                            }}
                        }};
                        el.chart.on('click', el.__sampleOrderStatisticsHandler);
                    }}, 200);
                """)

            def render_statistics(date_basis: str, count_basis: str) -> None:
                normalized_basis = "actual" if date_basis == "actual" else "planned"
                normalized_count_basis = "samples" if count_basis == "samples" else "orders"
                value_name = "按样品数" if normalized_count_basis == "samples" else "按订单数"
                statistics = get_sample_order_monthly_statistics(
                    all_records,
                    date_basis=normalized_basis,
                    count_basis=normalized_count_basis,
                )
                chart_options = build_sample_order_statistics_chart(
                    statistics,
                    include_incomplete=normalized_basis == "planned",
                    value_name=value_name,
                )
                chart_container.clear()
                with chart_container:
                    if normalized_basis == "actual":
                        ui.label(
                            "已完成订单按实际交货日期归属月份；未完成订单没有实际交货日期，不计入此口径。"
                            "按时或延期仍以实际交货日期是否晚于计划交货日期判定。"
                        ).classes("text-sm text-gray-500 shrink-0")
                    else:
                        ui.label(
                            "显示过去11个月、当前月及已有未来计划月份；订单按计划交货日期归属月份，"
                            "实际交货不晚于计划日期为按时完成，晚于计划日期为延期完成。"
                        ).classes("text-sm text-gray-500 shrink-0")
                    if normalized_count_basis == "samples":
                        ui.label("样品数按每张订单的申请数量累加。").classes("text-sm text-teal-700 shrink-0")
                    chart = ui.echart(chart_options).classes("w-full flex-grow min-h-0 cursor-pointer")

                    def show_clicked_details(event: Any) -> None:
                        event_args = event.args if isinstance(event.args, dict) else {}
                        open_statistics_details(
                            option_text(event_args.get("month")),
                            option_text(event_args.get("category")),
                            normalized_basis,
                            normalized_count_basis,
                        )

                    bind_statistics_chart_click(chart, show_clicked_details)

            render_statistics("planned", "orders")
            date_basis_toggle.on_value_change(
                lambda event: render_statistics(
                    option_text(event.value, "planned"),
                    option_text(count_basis_toggle.value, "orders"),
                )
            )
            count_basis_toggle.on_value_change(
                lambda event: render_statistics(
                    option_text(date_basis_toggle.value, "planned"),
                    option_text(event.value, "orders"),
                )
            )
        statistics_dialog.open()

    def open_sample_order_import_dialog() -> None:
        """打开Excel上传、预览和确认导入弹窗。"""
        if not can_edit_base:
            ui.notify("仅研发助理可以导入样品单", type="warning", position="bottom")
            return
        import_state: dict[str, Any] = {"preview": None}
        import_dialog.clear()
        with import_dialog, ui.card().classes("w-[900px] max-w-[95vw] max-h-[90vh] p-5"):
            with ui.row().classes("w-full items-center justify-between"):
                with ui.row().classes("items-center gap-2"):
                    ui.icon("upload_file", color="blue")
                    ui.label("导入样品单Excel").classes("text-xl font-bold")
                ui.button(icon="close", on_click=import_dialog.close).props("flat round")
            ui.label("读取样品单基础信息、制样执行及首次/二次延期；物料、支援、公式提示和考核列不会导入。").classes(
                "text-sm text-gray-600"
            )
            ui.label(
                "上传后先预览，确认时一次性写入；每一行均按独立订单新增，制样负责人统一为叶子浩，且不会发送企业微信通知。"
            ).classes("text-sm text-orange-700")
            preview_container = ui.column().classes("w-full flex-grow overflow-y-auto gap-2")

            def render_import_preview(preview: SampleOrderImportPreview) -> None:
                preview_container.clear()
                with preview_container:
                    with ui.row().classes("w-full gap-3 flex-wrap"):
                        ui.badge(f"读取数据行：{preview.total_rows}", color="blue").props("outline")
                        ui.badge(f"可导入：{len(preview.records)}", color="green").props("outline")
                        ui.badge(f"异常行：{len(preview.errors)}", color="red").props("outline")
                        ui.badge(f"需留意：{len(preview.warnings)}", color="orange").props("outline")
                    if preview.records:
                        ui.label("前10条有效记录预览").classes("font-bold text-gray-700 mt-2")
                        for record in preview.records[:10]:
                            basic = record["basic_info"]
                            execution = record["execution"]
                            source_row = normalize_int(record["import_info"].get("source_row"), 0)
                            ui.label(
                                f"第{source_row}行 · {basic.get('sample_order_no', '')} · "
                                f"{basic.get('product_model', '')} · 数量{basic.get('application_qty', '')} · "
                                f"实际交样{execution.get('actual_delivery_date', '') or '未交样'} · "
                                f"延期{len(record['extensions'])}次"
                            ).classes("w-full text-sm text-gray-600 border-b pb-1")
                    if preview.errors:
                        with ui.expansion(
                            f"查看异常行（{len(preview.errors)}）",
                            icon="error_outline",
                        ).classes("w-full border rounded-lg mt-2"):
                            for error in preview.errors[:100]:
                                ui.label(error).classes("text-sm text-red-700")
                            if len(preview.errors) > 100:
                                ui.label("仅显示前100条异常信息").classes("text-xs text-gray-500")
                    if preview.warnings:
                        with ui.expansion(
                            f"查看需留意的数据（{len(preview.warnings)}）",
                            icon="warning_amber",
                        ).classes("w-full border rounded-lg mt-2"):
                            for warning_text in preview.warnings[:100]:
                                ui.label(warning_text).classes("text-sm text-orange-700")

            async def handle_excel_upload(event: UploadEventArguments) -> None:
                file_name = option_text(event.file.name)
                if not file_name.lower().endswith(".xlsx"):
                    ui.notify("请选择.xlsx格式的Excel文件", type="warning", position="bottom")
                    return
                try:
                    content = await event.file.read()
                    preview = parse_sample_order_excel(content, file_name)
                except Exception:
                    logger.exception("样品单Excel上传解析失败")
                    ui.notify("Excel解析失败，请检查文件格式", type="negative", position="bottom")
                    return
                import_state["preview"] = preview
                render_import_preview(preview)
                if preview.records:
                    ui.notify("Excel解析完成，请核对预览后确认导入", type="positive", position="bottom")
                else:
                    ui.notify("文件中没有可导入的有效记录", type="warning", position="bottom")

            ui.upload(
                label="选择样品单Excel文件",
                on_upload=handle_excel_upload,
                auto_upload=True,
                max_files=1,
            ).props('accept=".xlsx" max-file-size=20971520').classes("w-full")

            async def confirm_excel_import() -> None:
                preview = import_state.get("preview")
                if not isinstance(preview, SampleOrderImportPreview) or not preview.records:
                    ui.notify("请先上传并成功解析Excel文件", type="warning", position="bottom")
                    return
                result = await import_sample_order_records(
                    preview.records,
                    current_user,
                    current_role,
                    source_name=preview.source_name,
                )
                if not result.db_success:
                    ui.notify("导入写入失败，请稍后重试", type="negative", position="bottom")
                    return
                if result.code == "forbidden":
                    ui.notify("当前角色没有导入权限", type="negative", position="bottom")
                    return
                import_dialog.close()
                ui.notify(
                    f"导入完成：新增{result.imported_count}条",
                    type="positive" if result.imported_count else "info",
                    position="bottom",
                    timeout=6000,
                )
                refresh_dashboard()

            with ui.row().classes("w-full justify-end gap-2"):
                ui.button("取消", on_click=import_dialog.close).props("flat color=grey")
                ui.button("确认导入", icon="database", on_click=confirm_excel_import).props("color=primary")
        import_dialog.open()

    async def open_detail_dialog(target_record_id: Optional[str] = None) -> None:
        is_new = target_record_id is None
        if is_new and not can_edit_base:
            ui.notify("仅研发助理可以新建样品单", type="warning", position="bottom")
            return
        all_records = get_all_sample_order_records()
        if is_new:
            local_data = get_sample_order_template()
            local_data["record_id"] = uuid.uuid4().hex
        else:
            raw = all_records.get(target_record_id, {}) if isinstance(all_records, dict) else {}
            local_data = merge_with_sample_order_template(raw)
            if not local_data["record_id"]:
                ui.notify("未找到该样品单记录", type="warning", position="bottom")
                return

        detail_dialog.clear()
        with detail_dialog, ui.card().classes("w-full h-full rounded-none p-0"):
            preview_container = ui.row().classes("w-full gap-3 flex-wrap")
            preview_value_labels: list[Any] = []

            def render_preview() -> None:
                metrics = calculate_sample_order_metrics(local_data)
                preview_items = [
                    ("交样周期", f"{metrics['cycle_days']} 天" if metrics["cycle_days"] is not None else "--"),
                    ("提示信息", metrics["alert_message"]),
                    ("预期状况", metrics["expected_status"]),
                    (
                        "考核",
                        f"{metrics['assessment_days']} 天 / {metrics['assessment_score']} 分"
                        if metrics["assessment_days"] is not None
                        else "待实际交付",
                    ),
                ]
                if not preview_value_labels:
                    with preview_container:
                        for label, value in preview_items:
                            with ui.card().classes("min-w-40 flex-1 p-3 bg-slate-50 shadow-none border"):
                                ui.label(label).classes("text-xs text-gray-500")
                                preview_value_labels.append(ui.label(str(value)).classes("font-semibold text-gray-800"))
                    return
                for value_label, (_, value) in zip(preview_value_labels, preview_items):
                    value_label.set_text(str(value))

            def bind_text_input(
                label: str,
                target: dict,
                key: str,
                *,
                editable: bool,
                classes: str = "w-full",
                textarea: bool = False,
                refresh_metrics: bool = False,
                validate_person: bool = False,
            ) -> None:
                value = option_text(target.get(key))
                field = ui.textarea(label, value=value) if textarea else ui.input(label, value=value)
                field.props("outlined dense autogrow" if textarea else "outlined dense").classes(classes)
                if editable:

                    def set_value(event: Any, data: dict = target, data_key: str = key) -> None:
                        data[data_key] = option_text(event.value)
                        if refresh_metrics:
                            render_preview()

                    field.on_value_change(set_value)
                    if validate_person:

                        async def warn_unknown_name(
                            _event: Any = None,
                            label_text: str = label,
                            data: dict = target,
                            data_key: str = key,
                        ) -> None:
                            unknown_names = await find_unknown_wecom_names(data.get(data_key, ""))
                            if unknown_names:
                                display_label = label_text.rstrip(" *")
                                ui.notify(
                                    f"{display_label}未在企业微信通讯录中找到："
                                    f"{'、'.join(unknown_names)}，请检查是否有错别字",
                                    type="warning",
                                    position="bottom",
                                    multi_line=True,
                                )

                        field.on("blur", warn_unknown_name)
                else:
                    field.props("disable")

            def bind_date_input(
                label: str,
                target: dict,
                key: str,
                *,
                editable: bool,
                classes: str = "w-full",
                min_date: str = "",
                max_date: str = "",
            ) -> None:
                props = "outlined dense type=date"
                if min_date:
                    props += f" min={min_date}"
                if max_date:
                    props += f" max={max_date}"
                field = ui.input(label, value=option_text(target.get(key))).props(props).classes(classes)
                if editable:

                    def set_date(event: Any, data: dict = target, data_key: str = key) -> None:
                        data[data_key] = option_text(event.value)
                        render_preview()

                    field.on_value_change(set_date)
                else:
                    field.props("disable")

            basic = local_data["basic_info"]
            execution = local_data["execution"]
            extensions = local_data["extensions"]
            special_status = local_data["special_status"]
            delay_nature = local_data["delay_nature"]
            execution_editable = can_edit_base

            with ui.row().classes("w-full items-center justify-between px-6 py-3 bg-blue-600 text-white"):
                with ui.row().classes("items-center gap-3"):
                    ui.icon("fact_check", size="md")
                    ui.label("新增样品单" if is_new else f"样品单 {basic['sample_order_no']}").classes(
                        "text-xl font-bold"
                    )
                ui.button(icon="close", on_click=detail_dialog.close).props("flat round color=white")

            with ui.element("div").classes(
                "w-full flex-grow overflow-y-auto p-6 grid grid-cols-1 xl:grid-cols-2 gap-5 items-start"
            ):
                render_preview()
                with ui.card().classes("w-full h-full p-5 shadow-sm border bg-amber-50/50"):
                    with ui.row().classes("items-center gap-2 mb-3"):
                        ui.icon("description", color="blue")
                        ui.label("基础信息 · 研发助理维护").classes("text-lg font-bold")
                    with ui.grid().classes("w-full grid-cols-1 md:grid-cols-2 2xl:grid-cols-4 gap-4"):
                        bind_text_input("样品单号 *", basic, "sample_order_no", editable=can_edit_base)
                        bind_text_input("客户编码 *", basic, "customer_code", editable=can_edit_base)
                        bind_text_input("产品型号 *", basic, "product_model", editable=can_edit_base)
                        qty_field = (
                            ui.number(
                                "申请数量 *",
                                value=normalize_int(basic.get("application_qty"), 1),
                                min=1,
                                precision=0,
                            )
                            .props("outlined dense")
                            .classes("w-full")
                        )
                        if can_edit_base:

                            def set_qty(event: Any) -> None:
                                basic["application_qty"] = max(1, normalize_int(event.value, 1))

                            qty_field.on_value_change(set_qty)
                        else:
                            qty_field.props("disable")
                        bind_date_input("申请日期 *", basic, "application_date", editable=can_edit_base)
                        bind_text_input(
                            "申请人 *",
                            basic,
                            "applicant",
                            editable=can_edit_base,
                            validate_person=True,
                        )
                        bind_date_input("计划交货日期 *", basic, "planned_delivery_date", editable=can_edit_base)
                        bind_text_input(
                            "制样负责人",
                            execution,
                            "sample_owner",
                            editable=execution_editable,
                            validate_person=True,
                        )
                        bind_date_input(
                            "实际交货日期",
                            execution,
                            "actual_delivery_date",
                            editable=execution_editable,
                            max_date=date.today().isoformat(),
                        )
                        with ui.element("div").classes("w-full rounded-lg bg-blue-50 border border-blue-200 p-3"):
                            ui.label(f"系统提前 {SAMPLE_ORDER_WARNING_DAYS} 个工作日警示").classes(
                                "text-xs text-blue-600"
                            )

                    bind_text_input(
                        "备注",
                        basic,
                        "remark",
                        editable=can_edit_base,
                        classes="w-full",
                        textarea=True,
                    )

                with ui.card().classes("w-full h-full p-5 shadow-sm border bg-blue-100/50"):
                    with ui.row().classes("w-full items-center justify-between mb-1"):
                        with ui.row().classes("items-center gap-2"):
                            ui.icon("event_repeat", color="orange")
                            ui.label("延期历史 · 研发样品组长追加").classes("text-lg font-bold")
                        if can_edit_delay and not is_new:
                            ui.button("新增一次延期", icon="add", on_click=lambda: add_extension()).props(
                                "outline color=orange"
                            )

                    extension_container = ui.column().classes("w-full gap-3")

                    def remove_extension(index: int) -> None:
                        if 0 <= index < len(extensions) and not extensions[index].get("extension_id"):
                            extensions.pop(index)
                            render_extensions()
                            render_preview()

                    def add_extension() -> None:
                        extensions.append(normalize_extension({}))
                        render_extensions()
                        render_preview()

                    def render_extensions() -> None:
                        extension_container.clear()
                        with extension_container:
                            if not extensions:
                                ui.label("暂无延期记录").classes("text-sm text-gray-400 py-2")
                                return
                            for index, extension in enumerate(extensions):
                                saved = bool(extension.get("extension_id"))
                                with ui.element("div").classes(
                                    "w-full rounded-lg border p-4 "
                                    + ("bg-gray-10" if saved else "bg-orange-50 border-orange-300")
                                ):
                                    with ui.row().classes("w-full items-center justify-between mb-2"):
                                        ui.label(f"第 {index + 1} 次延期").classes("font-bold text-gray-800")
                                        if saved:
                                            ui.label(
                                                f"{extension.get('created_at', '')} · {extension.get('created_by', '')}"
                                            ).classes("text-xs text-gray-500")
                                        else:
                                            ui.button(
                                                icon="delete",
                                                on_click=lambda _=None, idx=index: remove_extension(idx),
                                            ).props("flat round dense color=negative")
                                    with ui.grid().classes("w-full grid-cols-1 md:grid-cols-12 gap-4"):
                                        bind_date_input(
                                            f"第{index + 1}次延期目标日期",
                                            extension,
                                            "target_date",
                                            editable=can_edit_delay and not saved,
                                            classes="w-full md:col-span-3",
                                            min_date=date.today().isoformat() if not saved else "",
                                        )
                                        bind_text_input(
                                            f"第{index + 1}次延期原因",
                                            extension,
                                            "reason",
                                            editable=can_edit_delay and not saved,
                                            classes="w-full md:col-span-9",
                                            refresh_metrics=True,
                                        )

                    render_extensions()

                    with ui.row().classes("items-center gap-2 mb-1"):
                        ui.icon("flag_circle", color="purple")
                        ui.label("订单特殊状态 · 研发样品组长设置").classes("text-lg font-bold")
                    ui.label("暂停、作废及恢复正常都会通知申请人和研发经理。").classes("text-sm text-purple-700 mb-3")
                    with ui.grid().classes("w-full grid-cols-1 md:grid-cols-2 xl:grid-cols-4 gap-4"):
                        status_select = (
                            ui.select(
                                SAMPLE_ORDER_SPECIAL_STATUSES,
                                label="订单特殊状态",
                                value=option_text(special_status.get("status"), "正常"),
                            )
                            .props("outlined dense")
                            .classes("w-full")
                        )
                        if can_edit_special_status and not is_new:

                            def set_special_status(event: Any) -> None:
                                special_status["status"] = option_text_in(
                                    event.value,
                                    SAMPLE_ORDER_SPECIAL_STATUSES,
                                    "正常",
                                )
                                render_preview()

                            status_select.on_value_change(set_special_status)
                        else:
                            status_select.props("disable")
                        bind_text_input(
                            "状态设置原因",
                            special_status,
                            "reason",
                            editable=can_edit_special_status and not is_new,
                        )
                    status_history = special_status.get("history", [])
                    if isinstance(status_history, list) and status_history:
                        with ui.expansion("特殊状态变更记录", icon="history").classes("w-full mt-3 border rounded-lg"):
                            for history_item in reversed(status_history):
                                if isinstance(history_item, dict):
                                    ui.label(
                                        f"{history_item.get('updated_at', '')} · "
                                        f"{history_item.get('old_status', '')} → {history_item.get('status', '')} · "
                                        f"{history_item.get('reason', '')} · {history_item.get('updated_by', '')}"
                                    ).classes("text-sm text-gray-600")

                with ui.card().classes("w-full p-5 shadow-sm border xl:col-span-2 bg-red-100/50"):
                    with ui.row().classes("w-full items-center justify-between mb-1"):
                        with ui.row().classes("items-center gap-2"):
                            ui.icon("sell", color="red")
                            ui.label("延期性质标记 · 研发经理维护").classes("text-lg font-bold")
                        current_nature_tag = option_text(delay_nature.get("tag"))
                        if current_nature_tag:
                            ui.badge(current_nature_tag, color="green").props("outline")
                        elif is_delay_nature_pending(local_data):
                            ui.badge("待标记", color="red")
                    ui.label(
                        "订单实际交付后，对样品组长填写的原始延期原因补充标准性质标签；历史标签会优先列出。"
                    ).classes("text-sm text-red-700 mb-3")
                    nature_applicable = bool(
                        parse_iso_date(execution.get("actual_delivery_date"))
                        and extensions
                        and special_status.get("status") != "作废"
                    )
                    nature_state = {"selected": current_nature_tag, "custom": ""}
                    if not nature_applicable:
                        ui.label("订单完成且存在延期记录后才需要标记。").classes("text-sm text-gray-500")
                    elif can_mark_delay_nature:
                        nature_catalog = get_delay_nature_catalog(all_records)
                        with ui.grid().classes("w-full grid-cols-1 md:grid-cols-4 xl:grid-cols-6 gap-4"):
                            nature_select = (
                                ui.select(
                                    nature_catalog,
                                    label="优先选择已有标签",
                                    value=current_nature_tag or None,
                                )
                                .props("outlined dense clearable")
                                .classes("w-full")
                            )

                            def set_selected_nature(event: Any) -> None:
                                nature_state["selected"] = option_text(event.value)

                            nature_select.on_value_change(set_selected_nature)
                            new_nature_input = (
                                ui.input("没有合适标签时输入新标签")
                                .props("outlined dense maxlength=50")
                                .classes("w-full")
                            )

                            def set_custom_nature(event: Any) -> None:
                                nature_state["custom"] = option_text(event.value)

                            new_nature_input.on_value_change(set_custom_nature)

                        async def save_delay_nature() -> None:
                            tag = option_text(nature_state.get("custom")) or option_text(nature_state.get("selected"))
                            if not tag:
                                ui.notify("请选择已有标签或输入新标签", type="warning", position="bottom")
                                return
                            result = await mark_sample_order_delay_nature(
                                local_data["record_id"],
                                tag,
                                current_user,
                                current_role,
                                expected_revision=get_record_revision(local_data),
                            )
                            if result.changed:
                                detail_dialog.close()
                                ui.notify(f"延期性质已标记为：{tag}", type="positive", position="bottom")
                                refresh_dashboard()
                            elif result.code == "revision_conflict":
                                ui.notify("记录已被其他人更新，请关闭后重新打开", type="warning", position="bottom")
                            elif result.code == "no_change":
                                ui.notify("延期性质没有变化", type="info", position="bottom")
                            elif result.code == "nature_not_applicable":
                                ui.notify("该订单当前不符合延期性质标记条件", type="warning", position="bottom")
                            else:
                                ui.notify("延期性质保存失败，请稍后重试", type="negative", position="bottom")

                        with ui.row().classes("w-full justify-end mt-3"):
                            ui.button("保存性质标记", icon="sell", on_click=save_delay_nature).props("color=red")
                    else:
                        ui.label(f"当前标签：{current_nature_tag or '待研发经理标记'}").classes("text-sm text-gray-600")

                    nature_history = delay_nature.get("history", [])
                    if isinstance(nature_history, list) and nature_history:
                        with ui.expansion("性质标记历史", icon="history").classes("w-full mt-3 border rounded-lg"):
                            for history_item in reversed(nature_history):
                                if isinstance(history_item, dict):
                                    ui.label(
                                        f"{history_item.get('marked_at', '')} · "
                                        f"{history_item.get('old_tag', '') or '未标记'} → "
                                        f"{history_item.get('tag', '')} · {history_item.get('marked_by', '')}"
                                    ).classes("text-sm text-gray-600")

                if local_data.get("operation_log"):
                    with ui.expansion("操作记录", icon="history").classes("w-full border rounded-lg xl:col-span-2"):
                        for log_item in reversed(local_data["operation_log"][-20:]):
                            if isinstance(log_item, dict):
                                ui.label(
                                    f"{log_item.get('time', '')} · {log_item.get('user', '')}"
                                    f"（{log_item.get('role', '')}）· {log_item.get('action', '')}"
                                ).classes("text-sm text-gray-600")

            async def save_current_record() -> None:
                errors = validate_sample_order_submission(
                    local_data,
                    check_basic=can_edit_base,
                    check_execution=execution_editable,
                    check_delay=can_edit_delay,
                    check_special_status=can_edit_special_status,
                )
                if errors:
                    ui.notify("；".join(errors), type="warning", position="bottom", multi_line=True)
                    return
                result = await save_sample_order_record(
                    local_data,
                    current_user,
                    current_role,
                    is_new=is_new,
                )
                if result.changed:
                    detail_dialog.close()
                    ui.notify("样品单已保存", type="positive", position="bottom")
                    if result.notification_failures:
                        ui.notify(
                            "；".join(result.notification_failures),
                            type="warning",
                            position="bottom",
                            multi_line=True,
                        )
                    refresh_dashboard()
                elif result.code == "revision_conflict":
                    ui.notify("记录已被其他人更新，请关闭后重新打开", type="warning", position="bottom")
                elif result.code == "invalid" and isinstance(result.record, dict):
                    errors_value = result.record.get("errors", [])
                    ui.notify("；".join(str(item) for item in errors_value), type="warning", position="bottom")
                elif result.code == "forbidden":
                    ui.notify("当前角色没有此项编辑权限", type="negative", position="bottom")
                elif result.code == "extension_history_conflict":
                    ui.notify("延期历史已变化或被修改，请关闭后重新打开", type="warning", position="bottom")
                else:
                    ui.notify("保存失败，请稍后重试", type="negative", position="bottom")

            async def confirm_delete() -> None:
                result = await delete_sample_order_record(local_data["record_id"], current_role)
                confirm_dialog.close()
                if result.changed:
                    detail_dialog.close()
                    ui.notify("样品单已删除", type="positive", position="bottom")
                    refresh_dashboard()
                else:
                    ui.notify("删除失败或记录已不存在", type="negative", position="bottom")

            def open_delete_confirmation() -> None:
                confirm_dialog.clear()
                with confirm_dialog, ui.card().classes("w-96"):
                    ui.label("确认删除这张样品单？").classes("text-lg font-bold")
                    ui.label("删除后无法在页面中恢复。").classes("text-sm text-gray-500")
                    with ui.row().classes("w-full justify-end gap-2 mt-4"):
                        ui.button("取消", on_click=confirm_dialog.close).props("flat")
                        ui.button("确认删除", on_click=confirm_delete).props("color=negative")
                confirm_dialog.open()

            with ui.row().classes("w-full justify-end gap-2 px-6 py-3 border-t bg-white"):
                if can_delete and not is_new:
                    ui.button("删除", icon="delete", on_click=open_delete_confirmation).props("outline color=negative")
                ui.button("关闭", on_click=detail_dialog.close).props("outline color=grey")
                if can_edit_base or can_edit_delay or can_edit_special_status:
                    ui.button("保存", icon="save", on_click=save_current_record).props("color=primary")
        detail_dialog.open()

    with ui.header(elevated=True).classes("flex justify-between items-center bg-blue-500 h-12 px-4"):
        ui.image(f"{IMG_DIR}/Rayfine.png").classes("absolute w-20")
        ui.label("样品单执行看板").classes("text-white text-xl font-bold absolute left-1/2 transform -translate-x-1/2")
        with ui.avatar(size="lg").classes("cursor-pointer ml-auto -mt-3"):
            ui.image(current_display_path)
            with ui.menu().props("auto-close"):
                ui.menu_item(f"你好, {current_user}")
                ui.separator()
                ui.menu_item("返回主界面", on_click=lambda: ui.navigate.to("/main"))
                ui.separator()
                ui.menu_item("注销登录", on_click=lambda: logout())

    with ui.element("div").classes("fixed top-12 bottom-0 left-0 right-0 overflow-hidden bg-slate-50"):
        with ui.column().classes("w-full h-full p-4 gap-3"):
            kpi_container = ui.row().classes("w-full gap-3")

            def apply_filters() -> None:
                page_state["page"] = 1
                refresh_dashboard()

            def change_page(target_page: int) -> None:
                page_state["page"] = max(1, target_page)
                refresh_dashboard()

            with ui.row().classes("w-full justify-between items-center bg-white p-3 shadow-sm rounded-lg"):
                with ui.row().classes("gap-3 items-center"):
                    ui.input("搜索单号/客户/型号/申请人/负责人").props("dense outlined clearable").bind_value(
                        page_state, "search_keyword"
                    ).classes("w-80")
                    ui.select(FILTER_OPTIONS, label="状态筛选").props("dense outlined").bind_value(
                        page_state, "filter_state"
                    ).classes("w-40")
                    ui.button("查询", icon="search", on_click=apply_filters).props("outline color=primary")
                    ui.button("刷新", icon="refresh", on_click=lambda: refresh_dashboard()).props("flat color=primary")
                    ui.button(
                        "近12个月统计",
                        icon="stacked_bar_chart",
                        on_click=open_statistics_dialog,
                    ).props("flat color=primary")
                with ui.row().classes("items-center gap-3"):
                    ui.label("默认仅显示未交样订单，点击卡片查看详情").classes("text-xs text-gray-500")
                    if can_edit_base:
                        ui.button(
                            "导入Excel",
                            icon="upload_file",
                            on_click=open_sample_order_import_dialog,
                        ).props("outline color=primary")
                        ui.button("录入样品单", icon="add", on_click=lambda: open_detail_dialog()).props(
                            "color=primary"
                        )

            with ui.element("div").classes("w-full flex-grow overflow-y-auto overflow-x-hidden p-1"):
                list_container = ui.column().classes("w-full gap-3")
            pagination_container = ui.row().classes("w-full justify-center items-center gap-2")

            def refresh_dashboard() -> None:
                # 主动刷新时同步版本号，避免本客户端在5秒轮询时重复刷新同一版本。
                current_version = db_storage.get_item(SAMPLE_ORDER_VERSION_KEY, 0)
                page_state["last_version"] = current_version
                calculation_key = (current_version, date.today().isoformat())
                if dashboard_data_cache.get("key") != calculation_key:
                    all_records = get_all_sample_order_records()
                    stored_values = all_records.values() if isinstance(all_records, dict) else []
                    valid_records = [
                        merge_with_sample_order_template(item) for item in stored_values if isinstance(item, dict)
                    ]
                    records_with_metrics = [
                        (record, calculate_sample_order_metrics(record)) for record in valid_records
                    ]
                    dashboard_data_cache["key"] = calculation_key
                    dashboard_data_cache["records_with_metrics"] = records_with_metrics
                else:
                    cached_entries = dashboard_data_cache.get("records_with_metrics")
                    records_with_metrics = cached_entries if isinstance(cached_entries, list) else []
                    valid_records = [record for record, _metrics in records_with_metrics]
                keyword = option_text(page_state.get("search_keyword")).lower()
                filter_value = option_text(
                    page_state.get("filter_state"),
                    DEFAULT_SAMPLE_ORDER_FILTER,
                )
                visible_entries: list[tuple[dict, dict]] = []
                for record, metrics in records_with_metrics:
                    basic = record["basic_info"]
                    execution = record["execution"]
                    searchable = " ".join(
                        option_text(value)
                        for value in (
                            basic.get("sample_order_no"),
                            basic.get("customer_code"),
                            basic.get("product_model"),
                            basic.get("applicant"),
                            execution.get("sample_owner"),
                        )
                    ).lower()
                    if keyword and keyword not in searchable:
                        continue
                    if not sample_order_matches_filter(
                        record,
                        filter_value,
                        calculated_metrics=metrics,
                    ):
                        continue
                    visible_entries.append((record, metrics))

                visible_entries.sort(
                    key=lambda entry: (
                        entry[0]["execution"].get("actual_delivery_date", "") != "",
                        entry[1].get("effective_target_date", "9999-12-31"),
                        entry[0]["basic_info"].get("sample_order_no", ""),
                    )
                )
                total_visible = len(visible_entries)
                total_pages = max(
                    1,
                    (total_visible + SAMPLE_ORDER_CARD_PAGE_SIZE - 1) // SAMPLE_ORDER_CARD_PAGE_SIZE,
                )
                current_page = min(
                    max(1, normalize_int(page_state.get("page"), 1)),
                    total_pages,
                )
                page_state["page"] = current_page
                page_start = (current_page - 1) * SAMPLE_ORDER_CARD_PAGE_SIZE
                page_entries = visible_entries[page_start : page_start + SAMPLE_ORDER_CARD_PAGE_SIZE]

                list_container.clear()
                with list_container:
                    if not page_entries:
                        empty_text = (
                            "暂无未交样订单，可通过状态筛选查看已完成或全部订单"
                            if filter_value == DEFAULT_SAMPLE_ORDER_FILTER and not keyword
                            else "没有符合当前条件的样品单"
                        )
                        ui.label(empty_text).classes("text-gray-500 m-auto mt-10")
                    for record, metrics in page_entries:
                        row = build_sample_order_row(record, calculated_metrics=metrics)
                        level = option_text(row.get("attention_level"), "normal")
                        nature_pending = bool(row.get("nature_pending"))
                        many_delays = normalize_int(row.get("delay_count"), 0) > SAMPLE_ORDER_DELAY_ATTENTION_THRESHOLD
                        card_color_classes, border_color = get_sample_order_card_palette(
                            level,
                            nature_pending=nature_pending,
                            can_mark_delay_nature=can_mark_delay_nature,
                        )
                        if many_delays and not (nature_pending and can_mark_delay_nature):
                            border_color = "#a855f7"
                        card_classes = (
                            "sample-order-card w-full border border-l-4 rounded-md p-3 "
                            "cursor-pointer transition-all " + card_color_classes
                        )
                        with ui.element("div").classes(card_classes) as card:

                            async def open_card_detail(
                                _event: Any,
                                target_record_id: str = option_text(row.get("record_id")),
                            ) -> None:
                                if target_record_id:
                                    await open_detail_dialog(target_record_id)

                            card.style(f"border-left-color: {border_color}")
                            card.on("click", open_card_detail)
                            with ui.element("div").classes(
                                "grid w-full grid-cols-1 "
                                "lg:grid-cols-[minmax(230px,1fr)_minmax(0,1fr)_minmax(250px,auto)] "
                                "items-center gap-x-6 gap-y-2"
                            ):
                                with ui.column().classes("gap-1 min-w-0"):
                                    with ui.row().classes("items-center gap-2 flex-wrap"):
                                        ui.label(option_text(row.get("sample_order_no"), "未填写单号")).classes(
                                            "font-mono text-base text-gray-800"
                                        )
                                        expected_status = option_text(
                                            row.get("expected_status"),
                                            "待定",
                                        )
                                        status_color = {
                                            "按期": "green",
                                            "按计划": "green",
                                            "按当前目标": "blue",
                                            "延期": "red",
                                            "暂停": "purple",
                                            "作废": "grey",
                                        }.get(expected_status, "blue")
                                        ui.badge(expected_status, color=status_color).props("outline")
                                        if nature_pending:
                                            ui.badge("待性质标记", color="red").props("outline")
                                        if many_delays:
                                            ui.badge(
                                                f"已延期{row.get('delay_count', 0)}次",
                                                color="purple",
                                            ).props("outline")
                                    ui.label(option_text(row.get("alert_message"), "暂无提示")).classes(
                                        "text-base font-bold text-gray-600"
                                    )
                                    ui.label(f"申请人：{option_text(row.get('applicant'), '-')}").classes(
                                        "text-sm text-gray-500"
                                    )

                                with ui.column().classes("gap-1 min-w-0"):
                                    with ui.row().classes("w-full items-center gap-x-4 gap-y-1 flex-wrap"):
                                        ui.label(f"产品型号：{option_text(row.get('product_model'), '-')}").classes(
                                            "font-bold text-gray-800 text-base"
                                        )
                                        ui.label(f"客户编码：{option_text(row.get('customer_code'), '-')}").classes(
                                            "text-sm text-gray-600 whitespace-nowrap"
                                        )
                                        ui.label(f"数量：{row.get('application_qty', '-')}").classes(
                                            "text-sm text-gray-600 whitespace-nowrap"
                                        )
                                    planned_delivery = option_text(
                                        row.get("planned_delivery_date"),
                                        "-",
                                    )
                                    current_target = option_text(row.get("latest_extension_target")) or planned_delivery
                                    ui.label(f"计划交期：{planned_delivery} · 当前目标：{current_target}").classes(
                                        "text-sm text-gray-600"
                                    )
                                    latest_reason = option_text(row.get("latest_extension_reason"))
                                    if latest_reason:
                                        ui.label(f"最新延期原因：{latest_reason}").classes(
                                            "text-sm text-orange-700 line-clamp-1"
                                        )
                                    elif option_text(row.get("remark")):
                                        ui.label(f"备注：{option_text(row.get('remark'))}").classes(
                                            "text-sm text-gray-500 line-clamp-1"
                                        )

                                with ui.column().classes("gap-1 min-w-0 text-sm"):
                                    ui.label(f"制样负责人：{option_text(row.get('sample_owner'), '-')}").classes(
                                        "text-gray-700 whitespace-nowrap"
                                    )
                                    ui.label(
                                        f"申请日期：{option_text(row.get('application_date'), '-')} · "
                                        f"实际交样：{sample_order_delivery_display(row.get('actual_delivery_date'))}"
                                    ).classes("text-gray-500 whitespace-nowrap")
                                    nature_tag = option_text(row.get("delay_nature_tag"))
                                    if nature_tag:
                                        ui.label(f"延期性质：{nature_tag}").classes("text-red-700 whitespace-nowrap")
                                    assessment_score = row.get("assessment_score", "")
                                    if assessment_score != "":
                                        ui.label(
                                            f"考核：{row.get('assessment_days', '-')}天 / {assessment_score}分"
                                        ).classes("text-gray-500 whitespace-nowrap")

                pagination_container.clear()
                with pagination_container:
                    if total_visible > SAMPLE_ORDER_CARD_PAGE_SIZE:
                        previous_button = ui.button(
                            "上一页",
                            icon="chevron_left",
                            on_click=lambda _=None, target=current_page - 1: change_page(target),
                        ).props("flat color=primary")
                        if current_page <= 1:
                            previous_button.props("disable")
                        ui.label(f"第 {current_page}/{total_pages} 页 · 共 {total_visible} 条").classes(
                            "text-sm text-gray-500"
                        )
                        next_button = ui.button(
                            "下一页",
                            icon="chevron_right",
                            on_click=lambda _=None, target=current_page + 1: change_page(target),
                        ).props("flat color=primary icon-right")
                        if current_page >= total_pages:
                            next_button.props("disable")

                if page_state.get("kpi_cache_key") == calculation_key:
                    return

                metrics_list = [metrics for _, metrics in records_with_metrics]
                total_count = len(valid_records)
                scores = [
                    item["assessment_score"] for item in metrics_list if isinstance(item.get("assessment_score"), int)
                ]
                average_score = round(sum(scores) / len(scores), 1) if scores else "--"

                def kpi_records(label: str) -> list[dict]:
                    """收集统计卡片对应的订单，用于生成悬停明细。"""
                    return [
                        record
                        for record, metrics in records_with_metrics
                        if sample_order_matches_kpi(record, metrics, label)
                    ]

                in_progress_records = kpi_records("制样中")
                warning_records = kpi_records("预警")
                overdue_records = kpi_records("延期")
                many_delay_records = kpi_records("多次延期")
                kpi_items = [
                    ("全部样品单", total_count, "inventory_2", "blue", None),
                    ("制样中", len(in_progress_records), "pending_actions", "indigo", in_progress_records),
                    ("预警", len(warning_records), "notifications_active", "orange", warning_records),
                    ("延期", len(overdue_records), "warning", "red", overdue_records),
                    ("多次延期", len(many_delay_records), "repeat", "purple", many_delay_records),
                ]
                if can_mark_delay_nature:
                    pending_records = kpi_records("待性质标记")
                    kpi_items.append(("待性质标记", len(pending_records), "sell", "red", pending_records))
                if can_view_sample_order_average_score(current_role):
                    # score_records = kpi_records("平均考核分")
                    kpi_items.append(("平均考核分", average_score, "military_tech", "green", None))
                kpi_container.clear()
                with kpi_container:
                    for label, value, icon, color, detail_records in kpi_items:
                        with (
                            ui.card()
                            .classes("flex-1 min-w-40 p-3 shadow-sm border-l-4")
                            .style(f"border-left-color: var(--q-{color})")
                        ):
                            with ui.row().classes("w-full justify-between items-center"):
                                with ui.column().classes("gap-0"):
                                    ui.label(label).classes("text-xs text-gray-500")
                                    ui.label(str(value)).classes("text-2xl font-bold text-gray-800")
                                ui.icon(icon, color=color, size="md")
                            if detail_records is not None:
                                with ui.tooltip().classes(
                                    "bg-slate-800/80 text-white p-3 max-w-3xl max-h-[70vh] overflow-y-auto"
                                ):
                                    ui.label(f"{label}对应订单（{len(detail_records)}条）").classes("font-bold mb-2")
                                    if detail_records:
                                        detail_html = "".join(
                                            "<div class='break-inside-avoid whitespace-nowrap'>"
                                            f"{escape(option_text(detail_record['basic_info'].get('sample_order_no'), '未填写单号'))} · "
                                            f"{escape(option_text(detail_record['basic_info'].get('product_model'), '未填写型号'))} · "
                                            f"申请人：{escape(option_text(detail_record['basic_info'].get('applicant'), '未填写申请人'))}"
                                            "</div>"
                                            for detail_record in detail_records
                                        )
                                        ui.html(detail_html, sanitize=False).classes(
                                            "text-xs leading-5 columns-2 gap-5"
                                        )
                                    else:
                                        ui.label("暂无对应订单").classes("text-xs")
                page_state["kpi_cache_key"] = calculation_key

            async def refresh_if_version_changed() -> None:
                latest_version = await db_storage.get_fresh_item(SAMPLE_ORDER_VERSION_KEY, 0)
                if latest_version != page_state.get("last_version"):
                    await db_storage.refresh_json_entities(SAMPLE_ORDER_ENTITY_NAMESPACE)
                    page_state["last_version"] = latest_version
                    refresh_dashboard()

            ui.timer(0.1, refresh_dashboard, once=True)
            ui.timer(5.0, refresh_if_version_changed)
            if record_id:
                ui.timer(0.2, lambda: open_detail_dialog(record_id), once=True)
